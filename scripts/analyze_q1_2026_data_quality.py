from __future__ import annotations

import math
import hashlib
import json
import re
import warnings
from collections import Counter
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
import xlrd
from matplotlib.colors import BoundaryNorm, ListedColormap


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_ROOT = PROJECT_ROOT / "四诊仪数据汇总"
OUTPUT_DIR = PROJECT_ROOT / "datasets" / "q1_2026_quality"
PLOT_DIR = OUTPUT_DIR / "plots"
WORKBOOK_OUTPUT = OUTPUT_DIR / "q1_2026_data_quality.xlsx"
GENERATED_REPORT_OUTPUT = OUTPUT_DIR / "q1_2026_data_quality_generated.md"
LINKED_EXPORT_OUTPUT = OUTPUT_DIR / "q1_2026_checkin_linked.xlsx"

Q1_START = pd.Timestamp("2026-01-01 00:00:00")
Q2_START = pd.Timestamp("2026-04-01 00:00:00")
DATE_INDEX = pd.date_range(Q1_START.normalize(), Q2_START.normalize() - pd.Timedelta(days=1), freq="D")

SESSION_MERGE_MINUTES = 10
REQUIRED_MODALITY_THRESHOLD = 0.5
MIN_SLOT_GAP_MINUTES = 30

ZHONGKE_MODALITIES = ["ask", "pulse", "tongue", "face"]
YUSHENGTANG_MODALITIES = ["ask", "pulse", "tongue", "voice"]
SLOT_LABELS = ["早", "中", "晚"]

plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "Arial Unicode MS", "DejaVu Sans"]
plt.rcParams["axes.unicode_minus"] = False

NAME_CLEANUP_RE = re.compile(r"[\s·•\.\-—_・]+")
LEADING_TS_RE = re.compile(r"^(\d{14})")
CAID_RE = re.compile(r'"CaId"\s*:\s*(\d+)')
CANAME_RE = re.compile(r'"CaName"\s*:\s*"(.*?)"')
TREAT_RE = re.compile(r'"TreatNumber"\s*:\s*(\d+)')
MODALITY_LABELS = {
    "ask": "问诊",
    "pulse": "脉诊波形",
    "tongue": "舌诊图片",
    "face": "面诊图片",
    "voice": "wav",
}


@dataclass(frozen=True)
class CohortContext:
    names: list[str]
    name_set: set[str]
    normalized_to_name: dict[str, str]


def ensure_output_dirs() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    PLOT_DIR.mkdir(parents=True, exist_ok=True)


def clean_name(value: object) -> str:
    text = str(value or "").strip()
    return NAME_CLEANUP_RE.sub("", text)


def build_cohort_context() -> CohortContext:
    cohort_dir = DATA_ROOT / "中科四诊仪11.09-12.10"
    names = sorted(path.name for path in cohort_dir.iterdir() if path.is_dir())
    normalized_to_name = {clean_name(name): name for name in names}
    return CohortContext(names=names, name_set=set(names), normalized_to_name=normalized_to_name)


def resolve_cohort_name(raw_name: object, cohort: CohortContext) -> str | None:
    text = str(raw_name or "").strip()
    if not text:
        return None
    if text in cohort.name_set:
        return text
    normalized = clean_name(text)
    if normalized in cohort.normalized_to_name:
        return cohort.normalized_to_name[normalized]
    matches = []
    for cohort_name in cohort.names:
        cohort_normalized = clean_name(cohort_name)
        if cohort_normalized and (cohort_normalized in normalized or normalized in cohort_normalized):
            matches.append(cohort_name)
    if len(matches) == 1:
        return matches[0]
    return None


def coerce_id(value: object) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
        if abs(value) >= 1e12:
            return format(value, ".0f")
        return str(value)
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def parse_datetime_value(value: object) -> pd.Timestamp | None:
    if value is None or value == "":
        return None
    try:
        ts = pd.to_datetime(value)
    except Exception:
        return None
    if pd.isna(ts):
        return None
    return pd.Timestamp(ts)


def parse_leading_timestamp(name: str) -> pd.Timestamp | None:
    match = LEADING_TS_RE.match(name)
    if not match:
        return None
    try:
        return pd.to_datetime(match.group(1), format="%Y%m%d%H%M%S")
    except Exception:
        return None


def date_in_q1(ts: pd.Timestamp | None) -> bool:
    return ts is not None and Q1_START <= ts < Q2_START


def read_text_relaxed(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="ignore")


def load_json_relaxed(path: Path) -> object | None:
    if not path.exists():
        return None
    try:
        return json.loads(read_text_relaxed(path))
    except Exception:
        return None


def parse_internal_datetime(text: object) -> pd.Timestamp | None:
    raw = str(text or "").strip()
    if not raw or raw.startswith("0001-01-01"):
        return None
    try:
        ts = pd.to_datetime(raw, utc=True)
    except Exception:
        try:
            ts = pd.to_datetime(raw)
        except Exception:
            return None
    if pd.isna(ts):
        return None
    if getattr(ts, "tzinfo", None) is not None:
        ts = ts.tz_convert("Asia/Shanghai").tz_localize(None)
    return pd.Timestamp(ts)


def json_file_has_payload(path: Path) -> bool:
    if not path.exists():
        return False
    text = read_text_relaxed(path).strip()
    if not text or text in {"[]", "{}"}:
        return False
    return True


def try_parse_numeric_sequence(value: object) -> list[float]:
    if value is None or value == "":
        return []
    if isinstance(value, list):
        result: list[float] = []
        for item in value:
            try:
                number = float(item)
            except Exception:
                continue
            if math.isfinite(number):
                result.append(number)
        return result
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return []
        if text.startswith("[") and text.endswith("]"):
            try:
                parsed = json.loads(text)
            except Exception:
                return []
            return try_parse_numeric_sequence(parsed)
        return []
    return []


def normalize_vector(values: list[float]) -> list[float]:
    if not values:
        return []
    min_value = min(values)
    max_value = max(values)
    if not math.isfinite(min_value) or not math.isfinite(max_value):
        return []
    if abs(max_value - min_value) < 1e-12:
        return [0.0 for _ in values]
    return [(value - min_value) / (max_value - min_value) for value in values]


def downsample_vector(values: list[float], target_size: int = 128) -> list[float]:
    if not values:
        return []
    if len(values) <= target_size:
        return values
    result: list[float] = []
    last_index = len(values) - 1
    for idx in range(target_size):
        source_index = round(idx * last_index / max(target_size - 1, 1))
        result.append(values[source_index])
    return result


def stable_numeric_hash(values: list[float]) -> str:
    payload = ",".join(f"{value:.6f}" for value in values)
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()


def mean_manhattan_distance(left: list[float], right: list[float]) -> float | None:
    if not left or not right:
        return None
    size = min(len(left), len(right))
    if size == 0:
        return None
    distance = sum(abs(left[idx] - right[idx]) for idx in range(size)) / size
    return float(distance)


def classify_zhongke_modality(sheet_name: str) -> str | None:
    mapping = {
        "脉诊数据": "pulse",
        "舌诊数据": "tongue",
        "面诊数据": "face",
        "问诊汇总": "ask",
    }
    return mapping.get(sheet_name)


def parse_customer_archive_name_map(old_root: Path, cohort: CohortContext) -> dict[str, str]:
    caid_to_name: dict[str, str] = {}
    for path in old_root.rglob("customerArchive.json"):
        text = read_text_relaxed(path)
        id_match = CAID_RE.search(text)
        name_match = CANAME_RE.search(text)
        if not id_match or not name_match:
            continue
        canonical_name = resolve_cohort_name(name_match.group(1), cohort)
        if canonical_name:
            caid_to_name[id_match.group(1)] = canonical_name
    return caid_to_name


def get_yushengtang_roots() -> list[Path]:
    return sorted(
        path
        for path in DATA_ROOT.iterdir()
        if path.is_dir() and path.name.startswith("玉生堂四诊仪")
    )


def parse_zhongke_records(cohort: CohortContext) -> tuple[pd.DataFrame, pd.DataFrame]:
    modality_rows: list[dict[str, object]] = []
    file_scope_rows: list[dict[str, object]] = []

    warnings.filterwarnings("ignore", message="Workbook contains no default style")

    excel_paths = sorted(
        path
        for path in DATA_ROOT.rglob("*")
        if path.is_file() and path.suffix.lower() in {".xls", ".xlsx"} and path.relative_to(DATA_ROOT).parts[0].startswith("中科")
    )

    for path in excel_paths:
        rel_path = path.relative_to(DATA_ROOT)
        raw = path.read_bytes()
        head = raw[:8]
        file_hits = 0
        file_modalities: set[str] = set()
        min_dt: pd.Timestamp | None = None
        max_dt: pd.Timestamp | None = None

        try:
            if head.startswith(b"PK"):
                workbook = openpyxl.load_workbook(BytesIO(raw), read_only=True, data_only=True)
                try:
                    for worksheet in workbook.worksheets:
                        modality = classify_zhongke_modality(worksheet.title)
                        if modality is None or modality == "ask":
                            continue
                        seen_keys: set[tuple[str, str, str, pd.Timestamp]] = set()
                        for row in worksheet.iter_rows(min_row=4, values_only=True):
                            values = list(row)
                            if len(values) < 8:
                                continue
                            canonical_name = resolve_cohort_name(values[2], cohort)
                            collected_at = parse_datetime_value(values[6])
                            case_id = coerce_id(values[7])
                            if not canonical_name or not case_id or not date_in_q1(collected_at):
                                continue
                            key = (canonical_name, case_id, modality, collected_at)
                            if key in seen_keys:
                                continue
                            seen_keys.add(key)
                            modality_rows.append(
                                {
                                    "source_vendor": "zhongke",
                                    "record_origin": "workbook",
                                    "user_name": canonical_name,
                                    "source_visit_id": case_id,
                                    "collected_at": collected_at,
                                    "modality": modality,
                                    "source_path": str(rel_path),
                                }
                            )
                            file_hits += 1
                            file_modalities.add(modality)
                            min_dt = collected_at if min_dt is None else min(min_dt, collected_at)
                            max_dt = collected_at if max_dt is None else max(max_dt, collected_at)
                finally:
                    workbook.close()
            elif head == bytes.fromhex("D0CF11E0A1B11AE1"):
                workbook = xlrd.open_workbook(file_contents=raw, on_demand=True)
                try:
                    if "问诊汇总" in workbook.sheet_names():
                        sheet = workbook.sheet_by_name("问诊汇总")
                        modality = "ask"
                        seen_keys: set[tuple[str, str, str, pd.Timestamp]] = set()
                        for row_idx in range(1, sheet.nrows):
                            values = sheet.row_values(row_idx)
                            if len(values) < 8:
                                continue
                            canonical_name = resolve_cohort_name(values[2], cohort)
                            collected_at = parse_datetime_value(values[6])
                            case_id = coerce_id(values[7])
                            if not canonical_name or not case_id or not date_in_q1(collected_at):
                                continue
                            key = (canonical_name, case_id, modality, collected_at)
                            if key in seen_keys:
                                continue
                            seen_keys.add(key)
                            modality_rows.append(
                                {
                                    "source_vendor": "zhongke",
                                    "record_origin": "workbook",
                                    "user_name": canonical_name,
                                    "source_visit_id": case_id,
                                    "collected_at": collected_at,
                                    "modality": modality,
                                    "source_path": str(rel_path),
                                }
                            )
                            file_hits += 1
                            file_modalities.add(modality)
                            min_dt = collected_at if min_dt is None else min(min_dt, collected_at)
                            max_dt = collected_at if max_dt is None else max(max_dt, collected_at)
                finally:
                    workbook.release_resources()
        except Exception as exc:
            file_scope_rows.append(
                {
                    "source_vendor": "zhongke",
                    "source_path": str(rel_path),
                    "matched_records": 0,
                    "matched_modalities": "",
                    "min_collected_at": None,
                    "max_collected_at": None,
                    "notes": f"parse_error: {type(exc).__name__}",
                }
            )
            continue

        if file_hits:
            file_scope_rows.append(
                {
                    "source_vendor": "zhongke",
                    "source_path": str(rel_path),
                    "matched_records": file_hits,
                    "matched_modalities": ",".join(sorted(file_modalities)),
                    "min_collected_at": min_dt,
                    "max_collected_at": max_dt,
                    "notes": "",
                }
            )

    if not modality_rows:
        return pd.DataFrame(), pd.DataFrame(file_scope_rows)

    modality_df = pd.DataFrame(modality_rows).drop_duplicates(
        subset=["source_vendor", "user_name", "source_visit_id", "collected_at", "modality"]
    )

    visits = (
        modality_df.assign(flag=True)
        .pivot_table(
            index=["source_vendor", "user_name", "source_visit_id", "collected_at"],
            columns="modality",
            values="flag",
            aggfunc="max",
            fill_value=False,
        )
        .reset_index()
    )
    visits.columns.name = None
    for modality in ZHONGKE_MODALITIES:
        if modality not in visits.columns:
            visits[modality] = False
    visits["record_origin"] = "workbook"
    visits["source_batch"] = "zhongke_q1_scan"
    visits["has_case_dir"] = True
    visits["source_path_count"] = (
        modality_df.groupby(["source_vendor", "user_name", "source_visit_id", "collected_at"])["source_path"].transform("nunique")
    )

    file_scope_df = pd.DataFrame(file_scope_rows)
    return visits, file_scope_df


def extract_case_json_scalars(case_dir: Path) -> dict[str, str]:
    scalar_values: dict[str, str] = {}
    candidate_files = [
        case_dir / "customerArchive.json",
        case_dir / "dataAsk.json",
        case_dir / "dataBasic.json",
        case_dir / "pulse" / "dataPulse.json",
        case_dir / "tongue" / "dataTongue.json",
        case_dir / "voice" / "dataVoice.json",
        case_dir / "numberOrder" / "numberOrder.json",
    ]
    for path in candidate_files:
        if not path.exists():
            continue
        text = read_text_relaxed(path)
        if "CaId" not in scalar_values:
            match = CAID_RE.search(text)
            if match:
                scalar_values["CaId"] = match.group(1)
        if "CaName" not in scalar_values:
            match = CANAME_RE.search(text)
            if match:
                scalar_values["CaName"] = match.group(1)
        if "TreatNumber" not in scalar_values:
            match = TREAT_RE.search(text)
            if match:
                scalar_values["TreatNumber"] = match.group(1)
    return scalar_values


def inspect_yushengtang_case_dir(case_dir: Path) -> dict[str, object] | None:
    pulse_path = case_dir / "pulse" / "dataPulse.json"
    pulse_payload = load_json_relaxed(pulse_path)
    internal_timestamp = None
    if isinstance(pulse_payload, dict):
        internal_timestamp = parse_internal_datetime(pulse_payload.get("StartTime"))
    timestamp = internal_timestamp or parse_leading_timestamp(case_dir.name)
    if not date_in_q1(timestamp):
        return None

    scalar_values = extract_case_json_scalars(case_dir)
    visit_id = scalar_values.get("TreatNumber", case_dir.name)
    tongue_path = case_dir / "tongue" / "dataTongue.json"
    voice_path = case_dir / "voice" / "dataVoice.json"
    ask_path = case_dir / "dataAsk.json"

    return {
        "source_vendor": "yushengtang",
        "record_origin": "case_dir",
        "user_name": None,
        "source_visit_id": visit_id,
        "collected_at": timestamp,
        "ask": json_file_has_payload(ask_path),
        "pulse": json_file_has_payload(pulse_path),
        "tongue": json_file_has_payload(tongue_path),
        "voice": json_file_has_payload(voice_path),
        "basic": json_file_has_payload(case_dir / "dataBasic.json"),
        "western": json_file_has_payload(case_dir / "western" / "dataWestern.json"),
        "number_order": json_file_has_payload(case_dir / "numberOrder" / "numberOrder.json"),
        "pdf": any(child.suffix.lower() == ".pdf" for child in case_dir.iterdir() if child.is_file()),
        "has_case_dir": True,
        "source_batch": case_dir.parent.name,
        "source_path": str(case_dir.relative_to(DATA_ROOT)),
        "caid": scalar_values.get("CaId", ""),
        "raw_source_name": scalar_values.get("CaName", ""),
    }


def parse_yushengtang_records(cohort: CohortContext) -> tuple[pd.DataFrame, pd.DataFrame]:
    yushengtang_roots = get_yushengtang_roots()
    if not yushengtang_roots:
        return pd.DataFrame(), pd.DataFrame()

    caid_to_name: dict[str, str] = {}
    for root in yushengtang_roots:
        caid_to_name.update(parse_customer_archive_name_map(root, cohort))

    records: list[dict[str, object]] = []
    file_scope_rows: list[dict[str, object]] = []

    for source_root in yushengtang_roots:
        for case_dir in sorted(path for path in source_root.iterdir() if path.is_dir()):
            case_info = inspect_yushengtang_case_dir(case_dir)
            if case_info is None:
                continue

            canonical_name = None
            if case_info["raw_source_name"]:
                canonical_name = resolve_cohort_name(case_info["raw_source_name"], cohort)
            if canonical_name is None and case_info["caid"]:
                canonical_name = caid_to_name.get(case_info["caid"])
            if canonical_name is None:
                continue

            case_info["user_name"] = canonical_name
            records.append(case_info)
            file_scope_rows.append(
                {
                    "source_vendor": "yushengtang",
                    "source_path": case_info["source_path"],
                    "matched_records": 1,
                    "matched_modalities": ",".join(
                        modality
                        for modality in ["ask", "pulse", "tongue", "voice"]
                        if bool(case_info.get(modality, False))
                    ),
                    "min_collected_at": case_info["collected_at"],
                    "max_collected_at": case_info["collected_at"],
                    "notes": "case_dir",
                }
            )

    visits_from_dirs = pd.DataFrame(records)

    workbook_rows: list[dict[str, object]] = []
    for source_root in yushengtang_roots:
        for workbook_path in sorted(source_root.glob("*.xlsx")):
            workbook = openpyxl.load_workbook(BytesIO(workbook_path.read_bytes()), read_only=True, data_only=True)
            worksheet = workbook.worksheets[0]
            try:
                try:
                    worksheet.reset_dimensions()
                except Exception:
                    pass
                header_row = list(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)))
                if len(header_row) >= 25 and str(header_row[0]).strip() == "序号":
                    idx_visit_id, idx_time, idx_name, idx_pulse, idx_tongue_a, idx_tongue_b = 1, 3, 4, 16, 22, 23
                else:
                    idx_visit_id, idx_time, idx_name, idx_pulse, idx_tongue_a, idx_tongue_b = 0, 2, 3, 14, 20, 21

                matched_count = 0
                min_dt: pd.Timestamp | None = None
                max_dt: pd.Timestamp | None = None
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    values = list(row)
                    if len(values) <= max(idx_visit_id, idx_time, idx_name):
                        continue
                    canonical_name = resolve_cohort_name(values[idx_name], cohort)
                    collected_at = parse_datetime_value(values[idx_time])
                    visit_id = coerce_id(values[idx_visit_id])
                    if not canonical_name or not visit_id or not date_in_q1(collected_at):
                        continue
                    pulse_summary_present = idx_pulse < len(values) and bool(str(values[idx_pulse] or "").strip())
                    tongue_summary_present = (
                        (idx_tongue_a < len(values) and bool(str(values[idx_tongue_a] or "").strip()))
                        or (idx_tongue_b < len(values) and bool(str(values[idx_tongue_b] or "").strip()))
                    )
                    workbook_rows.append(
                        {
                            "source_vendor": "yushengtang",
                            "record_origin": "summary_workbook",
                            "user_name": canonical_name,
                            "source_visit_id": visit_id,
                            "collected_at": collected_at,
                            "ask": False,
                            "pulse": pulse_summary_present,
                            "tongue": tongue_summary_present,
                            "voice": False,
                            "basic": False,
                            "western": False,
                            "number_order": False,
                            "pdf": False,
                            "has_case_dir": False,
                            "source_batch": workbook_path.parent.name,
                            "source_path": str(workbook_path.relative_to(DATA_ROOT)),
                            "caid": "",
                            "raw_source_name": canonical_name,
                        }
                    )
                    matched_count += 1
                    min_dt = collected_at if min_dt is None else min(min_dt, collected_at)
                    max_dt = collected_at if max_dt is None else max(max_dt, collected_at)

                if matched_count:
                    file_scope_rows.append(
                        {
                            "source_vendor": "yushengtang",
                            "source_path": str(workbook_path.relative_to(DATA_ROOT)),
                            "matched_records": matched_count,
                            "matched_modalities": "summary_workbook",
                            "min_collected_at": min_dt,
                            "max_collected_at": max_dt,
                            "notes": "workbook_summary",
                        }
                    )
            finally:
                workbook.close()

    workbook_df = pd.DataFrame(workbook_rows)
    if visits_from_dirs.empty:
        visits = workbook_df.copy()
    elif workbook_df.empty:
        visits = visits_from_dirs.copy()
    else:
        workbook_only = workbook_df.loc[
            ~workbook_df["source_visit_id"].isin(set(visits_from_dirs["source_visit_id"]))
        ].copy()
        visits = pd.concat([visits_from_dirs, workbook_only], ignore_index=True, sort=False)

    if visits.empty:
        return pd.DataFrame(), pd.DataFrame(file_scope_rows)

    for column in ["ask", "pulse", "tongue", "voice", "basic", "western", "number_order", "pdf", "has_case_dir"]:
        if column not in visits.columns:
            visits[column] = False
    visits = visits.drop_duplicates(subset=["source_vendor", "user_name", "source_visit_id"])
    visits["collected_at"] = pd.to_datetime(visits["collected_at"])
    visits["source_path_count"] = 1

    return visits, pd.DataFrame(file_scope_rows)


def extract_yushengtang_pulse_feature(case_dir: Path) -> dict[str, object] | None:
    pulse_payload = load_json_relaxed(case_dir / "pulse" / "dataPulse.json")
    if not isinstance(pulse_payload, dict):
        return None

    combined: list[float] = []
    candidate_keys = ["SinglePluse", "CunShang", "Cun", "GuanMai", "Chi", "ChiXia"]
    for key in candidate_keys:
        combined.extend(try_parse_numeric_sequence(pulse_payload.get(key)))
    if len(combined) < 32:
        for value in pulse_payload.values():
            sequence = try_parse_numeric_sequence(value)
            if len(sequence) >= 32:
                combined.extend(sequence)
    if len(combined) < 32:
        return None

    normalized = normalize_vector(combined)
    sampled = downsample_vector(normalized, 128)
    if len(sampled) < 32:
        return None
    return {
        "numeric_signal_hash": stable_numeric_hash(sampled),
        "numeric_signal_vector": sampled,
        "numeric_signal_length": len(combined),
        "numeric_signal_source": "pulse_json",
    }


def read_csv_relaxed(path: Path) -> pd.DataFrame:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            return pd.read_csv(path, encoding=encoding, low_memory=False)
        except Exception:
            continue
    text = read_text_relaxed(path)
    return pd.read_csv(BytesIO(text.encode("utf-8")), encoding="utf-8", low_memory=False)


def extract_zhongke_pulse_feature_map(visit_ids: set[str]) -> dict[str, dict[str, object]]:
    if not visit_ids:
        return {}

    feature_map: dict[str, dict[str, object]] = {}
    pulse_csv_paths = sorted(
        path
        for path in DATA_ROOT.rglob("*脉诊数据*.csv")
        if path.is_file() and path.relative_to(DATA_ROOT).parts[0].startswith("中科")
    )
    position_order = {"左寸": 0, "左关": 1, "左尺": 2, "右寸": 3, "右关": 4, "右尺": 5}

    for path in pulse_csv_paths:
        try:
            frame = read_csv_relaxed(path)
        except Exception:
            continue
        if frame.empty:
            continue
        if "病历编号" not in frame.columns:
            continue
        frame["病历编号"] = frame["病历编号"].map(coerce_id)
        matched = frame.loc[frame["病历编号"].isin(visit_ids)].copy()
        if matched.empty:
            continue

        numeric_columns = [
            column
            for column in matched.columns
            if column != "病历编号" and pd.api.types.is_numeric_dtype(pd.to_numeric(matched[column], errors="coerce"))
        ]
        if not numeric_columns:
            candidate_columns = [column for column in matched.columns if column != "病历编号"]
            numeric_columns = []
            for column in candidate_columns:
                coerced = pd.to_numeric(matched[column], errors="coerce")
                if coerced.notna().any():
                    matched[column] = coerced
                    numeric_columns.append(column)

        if not numeric_columns:
            continue

        for visit_id, subset in matched.groupby("病历编号", sort=False):
            if visit_id in feature_map:
                continue
            subset = subset.copy()
            if "部位" in subset.columns:
                subset["_position_order"] = subset["部位"].map(lambda value: position_order.get(str(value).strip(), 99))
                subset = subset.sort_values(["_position_order", "部位"])
            vector: list[float] = []
            for _, row in subset.iterrows():
                for column in numeric_columns:
                    try:
                        number = float(row[column])
                    except Exception:
                        continue
                    if math.isfinite(number):
                        vector.append(number)
            if len(vector) < 12:
                continue
            normalized = normalize_vector(vector)
            sampled = downsample_vector(normalized, 128)
            if len(sampled) < 12:
                continue
            feature_map[visit_id] = {
                "numeric_signal_hash": stable_numeric_hash(sampled),
                "numeric_signal_vector": sampled,
                "numeric_signal_length": len(vector),
                "numeric_signal_source": "pulse_csv",
            }
    return feature_map


def detect_duplicate_numeric_visits(visits: pd.DataFrame) -> pd.DataFrame:
    if visits.empty:
        return visits.copy()

    visits = visits.copy()
    visits["duplicate_numeric_flag"] = False
    visits["duplicate_numeric_type"] = ""
    visits["duplicate_numeric_partner"] = ""
    visits["duplicate_numeric_distance"] = pd.NA
    visits["numeric_signal_hash"] = ""
    visits["numeric_signal_source"] = ""

    feature_records: list[dict[str, object]] = []
    yushengtang_subset = visits.loc[visits["source_vendor"] == "yushengtang"].copy()
    for _, row in yushengtang_subset.iterrows():
        case_dir = DATA_ROOT / str(row["source_path"])
        feature = extract_yushengtang_pulse_feature(case_dir)
        if feature is None:
            continue
        feature_records.append(
            {
                "source_vendor": "yushengtang",
                "user_name": row["user_name"],
                "source_visit_id": str(row["source_visit_id"]),
                **feature,
            }
        )

    zhongke_ids = set(visits.loc[visits["source_vendor"] == "zhongke", "source_visit_id"].astype(str))
    zhongke_feature_map = extract_zhongke_pulse_feature_map(zhongke_ids)
    for _, row in visits.loc[visits["source_vendor"] == "zhongke"].iterrows():
        visit_id = str(row["source_visit_id"])
        feature = zhongke_feature_map.get(visit_id)
        if feature is None:
            continue
        feature_records.append(
            {
                "source_vendor": "zhongke",
                "user_name": row["user_name"],
                "source_visit_id": visit_id,
                **feature,
            }
        )

    if not feature_records:
        return visits

    feature_df = pd.DataFrame(feature_records).drop_duplicates(subset=["source_vendor", "user_name", "source_visit_id"])
    if feature_df.empty:
        return visits

    visits = visits.merge(
        feature_df.drop(columns=["numeric_signal_vector"]),
        on=["source_vendor", "user_name", "source_visit_id"],
        how="left",
        suffixes=("", "_feature"),
    )
    visits["numeric_signal_hash"] = visits["numeric_signal_hash"].fillna("")
    visits["numeric_signal_source"] = visits["numeric_signal_source"].fillna("")

    key_columns = ["source_vendor", "user_name", "source_visit_id"]
    updates: dict[tuple[str, str, str], dict[str, object]] = {}

    for (_, user_name, numeric_hash), subset in feature_df.groupby(["source_vendor", "user_name", "numeric_signal_hash"], sort=False):
        if not numeric_hash or len(subset) < 2:
            continue
        visit_ids = sorted(subset["source_visit_id"].astype(str).tolist())
        for visit_id in visit_ids:
            key = (str(subset.iloc[0]["source_vendor"]), str(user_name), str(visit_id))
            partners = [partner for partner in visit_ids if partner != visit_id]
            updates[key] = {
                "duplicate_numeric_flag": True,
                "duplicate_numeric_type": "exact_hash",
                "duplicate_numeric_partner": ",".join(partners[:3]),
                "duplicate_numeric_distance": 0.0,
            }

    near_threshold = 0.015
    for (source_vendor, user_name), subset in feature_df.groupby(["source_vendor", "user_name"], sort=False):
        rows = subset.to_dict("records")
        for left_idx in range(len(rows)):
            for right_idx in range(left_idx + 1, len(rows)):
                left = rows[left_idx]
                right = rows[right_idx]
                if left["numeric_signal_hash"] == right["numeric_signal_hash"]:
                    continue
                distance = mean_manhattan_distance(
                    list(left["numeric_signal_vector"]),
                    list(right["numeric_signal_vector"]),
                )
                if distance is None or distance > near_threshold:
                    continue
                for current, partner in ((left, right), (right, left)):
                    key = (str(source_vendor), str(user_name), str(current["source_visit_id"]))
                    existing = updates.get(key)
                    if existing and existing.get("duplicate_numeric_type") == "exact_hash":
                        continue
                    if existing and existing.get("duplicate_numeric_distance") is not None:
                        try:
                            if float(existing["duplicate_numeric_distance"]) <= distance:
                                continue
                        except Exception:
                            pass
                    updates[key] = {
                        "duplicate_numeric_flag": True,
                        "duplicate_numeric_type": "near_manhattan",
                        "duplicate_numeric_partner": str(partner["source_visit_id"]),
                        "duplicate_numeric_distance": round(distance, 6),
                    }

    if not updates:
        return visits

    for key, payload in updates.items():
        mask = (
            (visits["source_vendor"].astype(str) == key[0])
            & (visits["user_name"].astype(str) == key[1])
            & (visits["source_visit_id"].astype(str) == key[2])
        )
        for column, value in payload.items():
            visits.loc[mask, column] = value
    visits["duplicate_numeric_flag"] = visits["duplicate_numeric_flag"].fillna(False).astype(bool)
    return visits


def determine_required_modalities(visits: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    if visits.empty:
        return pd.DataFrame(rows)

    source_to_modalities = {
        "zhongke": ZHONGKE_MODALITIES,
        "yushengtang": YUSHENGTANG_MODALITIES,
    }
    for source_vendor, modality_columns in source_to_modalities.items():
        subset = visits.loc[visits["source_vendor"] == source_vendor].copy()
        if subset.empty:
            continue
        for modality in modality_columns:
            presence_rate = float(subset[modality].fillna(False).astype(bool).mean())
            rows.append(
                {
                    "source_vendor": source_vendor,
                    "modality": modality,
                    "presence_rate": presence_rate,
                    "is_required": presence_rate >= REQUIRED_MODALITY_THRESHOLD,
                }
            )
    return pd.DataFrame(rows)


def apply_visit_quality_flags(visits: pd.DataFrame, modality_rules: pd.DataFrame) -> pd.DataFrame:
    if visits.empty:
        return visits.copy()

    required_lookup: dict[str, list[str]] = {}
    for source_vendor, subset in modality_rules.groupby("source_vendor"):
        required_lookup[source_vendor] = subset.loc[subset["is_required"], "modality"].tolist()

    visits = visits.copy()
    required_modalities_col: list[str] = []
    missing_required_col: list[str] = []
    is_complete_col: list[bool] = []

    for _, row in visits.iterrows():
        required = required_lookup.get(row["source_vendor"], [])
        missing = [modality for modality in required if not bool(row.get(modality, False))]
        required_modalities_col.append(",".join(required))
        missing_required_col.append(",".join(missing))
        is_complete_col.append(bool(required) and not missing)

    visits["required_modalities"] = required_modalities_col
    visits["missing_required_modalities"] = missing_required_col
    visits["is_complete_visit"] = is_complete_col
    visits["date"] = visits["collected_at"].dt.normalize()
    visits["month"] = visits["collected_at"].dt.to_period("M").astype(str)
    return visits


def cluster_sessions(visits: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    if visits.empty:
        return pd.DataFrame(), pd.DataFrame()

    session_rows: list[dict[str, object]] = []
    anomaly_rows: list[dict[str, object]] = []

    grouped = visits.sort_values(["user_name", "collected_at", "source_vendor"]).groupby(["user_name", "date"], sort=True)
    for (user_name, date_value), day_visits in grouped:
        current_cluster: list[dict[str, object]] = []
        last_ts: pd.Timestamp | None = None

        def flush_cluster(cluster: list[dict[str, object]]) -> None:
            if not cluster:
                return
            cluster_df = pd.DataFrame(cluster)
            start_ts = cluster_df["collected_at"].min()
            end_ts = cluster_df["collected_at"].max()
            complete_by_source = cluster_df.groupby("source_vendor")["is_complete_visit"].max().to_dict()
            visit_count_by_source = cluster_df.groupby("source_vendor")["source_visit_id"].nunique().to_dict()
            raw_visit_count = int(cluster_df["source_visit_id"].nunique())
            complete_visit_count = int(cluster_df.loc[cluster_df["is_complete_visit"], "source_visit_id"].nunique())
            incomplete_visit_count = raw_visit_count - complete_visit_count
            source_count = int(cluster_df["source_vendor"].nunique())
            cluster_id = f"{user_name}|{date_value.date()}|{start_ts.strftime('%H%M%S')}"
            all_complete_triplicate = raw_visit_count >= 3 and incomplete_visit_count == 0
            dual_device_triplicate = sum(count >= 3 for count in visit_count_by_source.values()) >= 2
            duplicate_numeric_flag = bool(cluster_df.get("duplicate_numeric_flag", pd.Series(dtype=bool)).fillna(False).any())
            duplicate_types = sorted(
                {
                    str(value)
                    for value in cluster_df.get("duplicate_numeric_type", pd.Series(dtype=object)).fillna("")
                    if str(value).strip()
                }
            )
            duplicate_partners = sorted(
                {
                    str(value)
                    for value in cluster_df.get("duplicate_numeric_partner", pd.Series(dtype=object)).fillna("")
                    if str(value).strip()
                }
            )
            cheat_triplicate_10m = bool(
                (end_ts - start_ts) <= pd.Timedelta(minutes=SESSION_MERGE_MINUTES)
                and (all_complete_triplicate or dual_device_triplicate)
            )
            if cheat_triplicate_10m:
                anomaly_rows.append(
                    {
                        "user_name": user_name,
                        "date": date_value,
                        "anomaly_type": "triplicate_within_10m",
                        "cluster_id": cluster_id,
                        "cluster_start": start_ts,
                        "cluster_end": end_ts,
                        "visit_count": raw_visit_count,
                        "complete_visit_count": complete_visit_count,
                        "incomplete_visit_count": incomplete_visit_count,
                        "source_vendors": ",".join(sorted(cluster_df["source_vendor"].unique())),
                        "source_visit_counts": ",".join(
                            f"{source}:{visit_count_by_source[source]}" for source in sorted(visit_count_by_source)
                        ),
                        "visit_ids": ",".join(cluster_df["source_visit_id"].astype(str).tolist()),
                        "anomaly_detail": "dual_device_triplicate" if dual_device_triplicate else "all_complete_triplicate",
                    }
                )
            if duplicate_numeric_flag:
                anomaly_rows.append(
                    {
                        "user_name": user_name,
                        "date": date_value,
                        "anomaly_type": "duplicate_numeric",
                        "cluster_id": cluster_id,
                        "cluster_start": start_ts,
                        "cluster_end": end_ts,
                        "visit_count": raw_visit_count,
                        "complete_visit_count": complete_visit_count,
                        "incomplete_visit_count": incomplete_visit_count,
                        "source_vendors": ",".join(sorted(cluster_df["source_vendor"].unique())),
                        "source_visit_counts": ",".join(
                            f"{source}:{visit_count_by_source[source]}" for source in sorted(visit_count_by_source)
                        ),
                        "visit_ids": ",".join(cluster_df["source_visit_id"].astype(str).tolist()),
                        "anomaly_detail": "|".join(part for part in [",".join(duplicate_types), ",".join(duplicate_partners)] if part),
                    }
                )
            session_rows.append(
                {
                    "user_name": user_name,
                    "date": date_value,
                    "cluster_id": cluster_id,
                    "cluster_start": start_ts,
                    "cluster_end": end_ts,
                    "visit_count": raw_visit_count,
                    "complete_visit_count": complete_visit_count,
                    "incomplete_visit_count": incomplete_visit_count,
                    "source_count": source_count,
                    "source_vendors": ",".join(sorted(cluster_df["source_vendor"].unique())),
                    "source_visit_counts": ",".join(
                        f"{source}:{visit_count_by_source[source]}" for source in sorted(visit_count_by_source)
                    ),
                    "visit_ids": ",".join(cluster_df["source_visit_id"].astype(str).tolist()),
                    "has_complete_visit": bool(cluster_df["is_complete_visit"].any()),
                    "all_visits_complete": bool(cluster_df["is_complete_visit"].all()),
                    "cheat_triplicate_10m": cheat_triplicate_10m,
                    "contains_zhongke": "zhongke" in complete_by_source,
                    "contains_yushengtang": "yushengtang" in complete_by_source,
                    "zhongke_complete": bool(complete_by_source.get("zhongke", False)),
                    "yushengtang_complete": bool(complete_by_source.get("yushengtang", False)),
                    "raw_sources": ",".join(sorted(cluster_df["record_origin"].unique())),
                    "duplicate_numeric_flag": duplicate_numeric_flag,
                    "duplicate_numeric_type": ",".join(duplicate_types),
                    "duplicate_numeric_partner": ",".join(duplicate_partners),
                }
            )

        for row in day_visits.to_dict("records"):
            if last_ts is None or row["collected_at"] - last_ts <= pd.Timedelta(minutes=SESSION_MERGE_MINUTES):
                current_cluster.append(row)
            else:
                flush_cluster(current_cluster)
                current_cluster = [row]
            last_ts = row["collected_at"]
        flush_cluster(current_cluster)

    return pd.DataFrame(session_rows), pd.DataFrame(anomaly_rows)


def compute_gap_threshold_minutes(sessions: pd.DataFrame) -> tuple[float, pd.DataFrame]:
    rows: list[dict[str, object]] = []
    if sessions.empty:
        return 0.0, pd.DataFrame(rows)

    for (user_name, date_value), day_sessions in sessions.groupby(["user_name", "date"], sort=True):
        day_sessions = day_sessions.sort_values("cluster_start").reset_index(drop=True)
        previous_session = None
        for _, session in day_sessions.iterrows():
            if previous_session is not None:
                gap_minutes = float((session["cluster_start"] - previous_session["cluster_start"]).total_seconds() / 60.0)
                rows.append(
                    {
                        "user_name": user_name,
                        "date": date_value,
                        "from_cluster_id": previous_session["cluster_id"],
                        "to_cluster_id": session["cluster_id"],
                        "gap_minutes": gap_minutes,
                        "from_suspicious": bool(previous_session["cheat_triplicate_10m"]),
                        "to_suspicious": bool(session["cheat_triplicate_10m"]),
                        "from_complete": bool(previous_session["has_complete_visit"]),
                        "to_complete": bool(session["has_complete_visit"]),
                    }
                )
            previous_session = session

    gap_df = pd.DataFrame(rows)
    if gap_df.empty:
        return 0.0, gap_df

    return float(MIN_SLOT_GAP_MINUTES), gap_df


def assign_day_slots(sessions: pd.DataFrame, min_slot_gap_minutes: float) -> tuple[pd.DataFrame, pd.DataFrame]:
    if sessions.empty:
        return pd.DataFrame(), pd.DataFrame()

    slot_rows: list[dict[str, object]] = []
    day_rows: list[dict[str, object]] = []
    gap_threshold = pd.Timedelta(minutes=min_slot_gap_minutes)

    for (user_name, date_value), day_sessions in sessions.groupby(["user_name", "date"], sort=True):
        day_sessions = day_sessions.sort_values("cluster_start").reset_index(drop=True)
        accepted_slot_count = 0
        display_slot_count = 0
        last_accepted_ts: pd.Timestamp | None = None
        extra_session_count = 0
        spacing_violation_count = 0
        triplicate_count = int(day_sessions["cheat_triplicate_10m"].sum())
        complete_slot_count = 0

        for _, session in day_sessions.iterrows():
            slot_label = ""
            display_slot_label = SLOT_LABELS[min(display_slot_count, len(SLOT_LABELS) - 1)]
            if display_slot_count < len(SLOT_LABELS) - 1:
                display_slot_count += 1
            else:
                display_slot_count = len(SLOT_LABELS) - 1
            status = "missing"
            reason = ""

            if session["cheat_triplicate_10m"]:
                status = "suspicious"
                reason = "triplicate_within_10m"
            elif bool(session.get("duplicate_numeric_flag", False)):
                status = "suspicious"
                reason = "duplicate_numeric"
            elif last_accepted_ts is not None and session["cluster_start"] - last_accepted_ts < gap_threshold:
                status = "invalid"
                reason = "gap_lt_dynamic_threshold"
                spacing_violation_count += 1
            elif accepted_slot_count >= len(SLOT_LABELS):
                status = "invalid"
                reason = "overflow_gt_3"
                extra_session_count += 1
            else:
                slot_label = SLOT_LABELS[accepted_slot_count]
                accepted_slot_count += 1
                last_accepted_ts = session["cluster_start"]
                if session["has_complete_visit"]:
                    status = "complete"
                    complete_slot_count += 1
                else:
                    status = "incomplete"
                    reason = "no_complete_visit_in_cluster"

            slot_rows.append(
                {
                    "user_name": user_name,
                    "date": date_value,
                    "cluster_id": session["cluster_id"],
                    "cluster_start": session["cluster_start"],
                    "source_vendors": session["source_vendors"],
                    "visit_ids": session["visit_ids"],
                    "slot_label": slot_label,
                    "display_slot_label": display_slot_label,
                    "slot_status": status,
                    "reason": reason,
                    "gap_threshold_minutes": min_slot_gap_minutes,
                    "has_complete_visit": bool(session["has_complete_visit"]),
                    "all_visits_complete": bool(session["all_visits_complete"]),
                    "contains_zhongke": bool(session["contains_zhongke"]),
                    "contains_yushengtang": bool(session["contains_yushengtang"]),
                    "zhongke_complete": bool(session["zhongke_complete"]),
                    "yushengtang_complete": bool(session["yushengtang_complete"]),
                    "cheat_triplicate_10m": bool(session["cheat_triplicate_10m"]),
                    "duplicate_numeric_flag": bool(session.get("duplicate_numeric_flag", False)),
                    "duplicate_numeric_type": str(session.get("duplicate_numeric_type", "")),
                    "duplicate_numeric_partner": str(session.get("duplicate_numeric_partner", "")),
                }
            )

        day_rows.append(
            {
                "user_name": user_name,
                "date": date_value,
                "sessions_total": int(len(day_sessions)),
                "accepted_slots": accepted_slot_count,
                "complete_slots": complete_slot_count,
                "triplicate_clusters": triplicate_count,
                "spacing_violations": spacing_violation_count,
                "extra_sessions": extra_session_count,
                "gap_threshold_minutes": min_slot_gap_minutes,
                "is_full_valid_day": complete_slot_count == len(SLOT_LABELS),
                "is_partial_valid_day": 0 < complete_slot_count < len(SLOT_LABELS),
            }
        )

    return pd.DataFrame(slot_rows), pd.DataFrame(day_rows)


def build_user_summary(day_df: pd.DataFrame, cohort: CohortContext) -> pd.DataFrame:
    total_days = len(DATE_INDEX)
    base = pd.DataFrame({"user_name": cohort.names})
    if day_df.empty:
        summary = base.assign(
            observed_days=0,
            full_valid_days=0,
            partial_valid_days=0,
            triplicate_clusters=0,
            spacing_violations=0,
            extra_sessions=0,
        )
    else:
        summary = (
            day_df.groupby("user_name")
            .agg(
                observed_days=("date", "nunique"),
                full_valid_days=("is_full_valid_day", "sum"),
                partial_valid_days=("is_partial_valid_day", "sum"),
                triplicate_clusters=("triplicate_clusters", "sum"),
                spacing_violations=("spacing_violations", "sum"),
                extra_sessions=("extra_sessions", "sum"),
            )
            .reset_index()
        )
        summary = base.merge(summary, on="user_name", how="left").fillna(0)

    for column in ["observed_days", "full_valid_days", "partial_valid_days", "triplicate_clusters", "spacing_violations", "extra_sessions", "missing_days"]:
        if column == "missing_days":
            continue
        summary[column] = summary[column].astype(int)
    summary["missing_days"] = total_days - summary["full_valid_days"]
    summary["coverage_rate"] = summary["full_valid_days"] / total_days
    summary["missing_days"] = summary["missing_days"].astype(int)
    return summary


def build_source_summary(visits: pd.DataFrame, modality_rules: pd.DataFrame) -> pd.DataFrame:
    if visits.empty:
        return pd.DataFrame()
    rows: list[dict[str, object]] = []
    for source_vendor, subset in visits.groupby("source_vendor"):
        required = modality_rules.loc[
            (modality_rules["source_vendor"] == source_vendor) & (modality_rules["is_required"]),
            "modality",
        ].tolist()
        rows.append(
            {
                "source_vendor": source_vendor,
                "visit_count": int(len(subset)),
                "user_count": int(subset["user_name"].nunique()),
                "complete_visit_count": int(subset["is_complete_visit"].sum()),
                "incomplete_visit_count": int((~subset["is_complete_visit"]).sum()),
                "case_dir_missing_count": int((~subset["has_case_dir"]).sum()),
                "required_modalities": ",".join(required),
            }
        )
    return pd.DataFrame(rows)


def build_gap_summary(gap_df: pd.DataFrame, gap_threshold_minutes: float) -> pd.DataFrame:
    if gap_df.empty:
        return pd.DataFrame()

    eligible_mask = (
        (~gap_df["from_suspicious"])
        & (~gap_df["to_suspicious"])
        & gap_df["from_complete"]
        & gap_df["to_complete"]
        & (gap_df["gap_minutes"] > 0)
    )
    eligible = gap_df.loc[eligible_mask, "gap_minutes"]
    if eligible.empty:
        eligible = gap_df.loc[gap_df["gap_minutes"] > 0, "gap_minutes"]
    if eligible.empty:
        return pd.DataFrame()

    summary_rows = [
        {"metric": "gap_threshold_minutes", "value": round(gap_threshold_minutes, 2)},
        {"metric": "eligible_gap_count", "value": int(len(eligible))},
        {"metric": "p10_gap_minutes", "value": round(float(eligible.quantile(0.10)), 2)},
        {"metric": "p25_gap_minutes", "value": round(float(eligible.quantile(0.25)), 2)},
        {"metric": "median_gap_minutes", "value": round(float(eligible.quantile(0.50)), 2)},
        {"metric": "p75_gap_minutes", "value": round(float(eligible.quantile(0.75)), 2)},
        {"metric": "p90_gap_minutes", "value": round(float(eligible.quantile(0.90)), 2)},
        {"metric": "mean_gap_minutes", "value": round(float(eligible.mean()), 2)},
        {"metric": "valid_ratio_at_threshold", "value": round(float((eligible >= gap_threshold_minutes).mean()), 4)},
    ]
    return pd.DataFrame(summary_rows)


def split_visit_ids(value: object) -> list[str]:
    return [part.strip() for part in str(value or "").split(",") if part.strip()]


def format_missing_modalities(cluster_visits: pd.DataFrame) -> str:
    missing_modalities: list[str] = []
    for value in cluster_visits.get("missing_required_modalities", pd.Series(dtype=object)).fillna(""):
        missing_modalities.extend(split_visit_ids(str(value).replace("，", ",")))
    unique_modalities: list[str] = []
    for modality in missing_modalities:
        if modality and modality not in unique_modalities:
            unique_modalities.append(modality)
    if not unique_modalities:
        return ""
    return "、".join(MODALITY_LABELS.get(modality, modality) for modality in unique_modalities)


def build_detail_remark(slot_row: pd.Series, session_row: pd.Series, cluster_visits: pd.DataFrame) -> str:
    status = str(slot_row["slot_status"])
    missing_detail = format_missing_modalities(cluster_visits)
    duplicate_note = ""
    if bool(slot_row.get("duplicate_numeric_flag", False)):
        duplicate_type = str(slot_row.get("duplicate_numeric_type", "")).strip()
        duplicate_partner = str(slot_row.get("duplicate_numeric_partner", "")).strip()
        duplicate_label = "数值指纹一致" if "exact_hash" in duplicate_type else "数值高度相似"
        duplicate_note = f"疑似重复：{duplicate_label}"
        if duplicate_partner:
            duplicate_note = f"{duplicate_note}（关联记录 {duplicate_partner}）"
    if status == "complete":
        return duplicate_note
    if status == "incomplete":
        base = f"不完整：缺少{missing_detail}" if missing_detail else "不完整：缺少必需模态"
        return "；".join(part for part in [base, duplicate_note] if part)
    if status == "invalid":
        if slot_row["reason"] == "gap_lt_dynamic_threshold":
            base = f"间隔过短：小于 {float(slot_row['gap_threshold_minutes']):.2f} 分钟"
            return "；".join(part for part in [base, duplicate_note] if part)
        if slot_row["reason"] == "overflow_gt_3":
            base = "超出单日三次显示上限"
            return "；".join(part for part in [base, duplicate_note] if part)
        return "；".join(part for part in ["无效时段", duplicate_note] if part)
    if status == "suspicious":
        if slot_row["reason"] == "duplicate_numeric":
            return duplicate_note or "疑似重复：数值指纹重复"
        if int(session_row.get("incomplete_visit_count", 0)) > 0:
            if missing_detail:
                base = f"疑似作弊：两设备连续多次，且含补打；缺少{missing_detail}"
            else:
                base = "疑似作弊：两设备连续多次，且含补打"
            return "；".join(part for part in [base, duplicate_note] if part)
        base = "疑似作弊：10分钟内连续多次完整打卡"
        return "；".join(part for part in [base, duplicate_note] if part)
    return ""


def build_detail_export(day_slots: pd.DataFrame, sessions: pd.DataFrame, visits: pd.DataFrame) -> pd.DataFrame:
    if day_slots.empty:
        return pd.DataFrame()

    sessions_lookup = sessions.set_index("cluster_id")
    visits = visits.copy()
    visits["source_visit_id"] = visits["source_visit_id"].astype(str)
    rows: list[dict[str, object]] = []

    def source_has(cluster_visits: pd.DataFrame, source_vendor: str, column: str) -> bool:
        subset = cluster_visits.loc[cluster_visits["source_vendor"] == source_vendor]
        if subset.empty or column not in subset.columns:
            return False
        return bool(subset[column].fillna(False).astype(bool).any())

    for _, slot_row in day_slots.iterrows():
        session_row = sessions_lookup.loc[slot_row["cluster_id"]]
        visit_ids = split_visit_ids(session_row["visit_ids"])
        cluster_visits = visits.loc[
            (visits["user_name"] == slot_row["user_name"]) & (visits["source_visit_id"].isin(visit_ids))
        ].copy()

        rows.append(
            {
                "姓名": slot_row["user_name"],
                "日期": pd.Timestamp(slot_row["date"]).normalize(),
                "具体时间": pd.Timestamp(slot_row["cluster_start"]).strftime("%H:%M:%S"),
                "时段": slot_row["display_slot_label"],
                "状态": slot_row["slot_status"],
                "备注": build_detail_remark(slot_row, session_row, cluster_visits),
                "中科_问诊": "1" if source_has(cluster_visits, "zhongke", "ask") else "",
                "中科_脉诊波形": "1" if source_has(cluster_visits, "zhongke", "pulse") else "",
                "中科_舌诊图片": "1" if source_has(cluster_visits, "zhongke", "tongue") else "",
                "中科_面诊图片": "1" if source_has(cluster_visits, "zhongke", "face") else "",
                "中科_wav": "1" if source_has(cluster_visits, "zhongke", "voice") else "",
                "玉生堂_问诊": "1" if source_has(cluster_visits, "yushengtang", "ask") else "",
                "玉生堂_脉诊波形": "1" if source_has(cluster_visits, "yushengtang", "pulse") else "",
                "玉生堂_舌诊图片": "1" if source_has(cluster_visits, "yushengtang", "tongue") else "",
                "玉生堂_面诊图片": "",
                "玉生堂_wav": "1" if source_has(cluster_visits, "yushengtang", "voice") else "",
                "疑似重复数值": "1" if bool(slot_row.get("duplicate_numeric_flag", False)) else "",
                "cluster_id": slot_row["cluster_id"],
            }
        )

    detail_df = pd.DataFrame(rows).sort_values(["姓名", "日期", "具体时间", "时段"]).reset_index(drop=True)
    return detail_df


def write_linked_export(cohort: CohortContext, day_slots: pd.DataFrame, detail_df: pd.DataFrame) -> None:
    with pd.ExcelWriter(LINKED_EXPORT_OUTPUT, engine="xlsxwriter", datetime_format="yyyy-mm-dd") as writer:
        workbook = writer.book
        matrix_sheet = workbook.add_worksheet("打卡矩阵")
        detail_sheet = workbook.add_worksheet("详细记录")
        writer.sheets["打卡矩阵"] = matrix_sheet
        writer.sheets["详细记录"] = detail_sheet

        header_fmt = workbook.add_format({"bold": True, "bg_color": "#D9EAF7", "border": 1, "align": "center", "valign": "vcenter"})
        row_label_fmt = workbook.add_format({"bold": True, "border": 1, "align": "left", "valign": "vcenter"})
        blank_fmt = workbook.add_format({"border": 1, "align": "center", "valign": "vcenter"})
        green_link_fmt = workbook.add_format({"border": 1, "align": "center", "valign": "vcenter", "bg_color": "#C6EFCE", "font_color": "#0563C1", "underline": 1})
        yellow_link_fmt = workbook.add_format({"border": 1, "align": "center", "valign": "vcenter", "bg_color": "#FFEB9C", "font_color": "#0563C1", "underline": 1})
        orange_link_fmt = workbook.add_format({"border": 1, "align": "center", "valign": "vcenter", "bg_color": "#F4B183", "font_color": "#0563C1", "underline": 1})
        red_link_fmt = workbook.add_format({"border": 1, "align": "center", "valign": "vcenter", "bg_color": "#FFC7CE", "font_color": "#0563C1", "underline": 1})

        matrix_dates = [day.strftime("%Y-%m-%d") for day in DATE_INDEX]
        matrix_sheet.write(0, 0, "用户时段", header_fmt)
        for col_idx, date_label in enumerate(matrix_dates, start=1):
            matrix_sheet.write(0, col_idx, date_label, header_fmt)

        detail_columns = [column for column in detail_df.columns if column != "cluster_id"]
        for col_idx, column in enumerate(detail_columns):
            detail_sheet.write(0, col_idx, column, header_fmt)
        for row_idx, row in detail_df.iterrows():
            excel_row = row_idx + 1
            for col_idx, column in enumerate(detail_columns):
                detail_sheet.write(excel_row, col_idx, row[column], blank_fmt)

        status_priority = {"suspicious": 4, "invalid": 3, "incomplete": 2, "complete": 1}
        key_to_status: dict[tuple[str, str, str], str] = {}
        key_to_detail_row: dict[tuple[str, str, str], int] = {}
        key_to_device_count: dict[tuple[str, str, str], int] = {}
        key_to_duplicate_flag: dict[tuple[str, str, str], bool] = {}
        zhongke_columns = ["中科_问诊", "中科_脉诊波形", "中科_舌诊图片", "中科_面诊图片", "中科_wav"]
        yushengtang_columns = ["玉生堂_问诊", "玉生堂_脉诊波形", "玉生堂_舌诊图片", "玉生堂_面诊图片", "玉生堂_wav"]
        for row_idx, row in detail_df.iterrows():
            key = (str(row["姓名"]), str(row["时段"]), pd.Timestamp(row["日期"]).strftime("%Y-%m-%d"))
            status = str(row["状态"])
            device_count = 0
            if any(str(row[column] or "").strip() for column in zhongke_columns):
                device_count += 1
            if any(str(row[column] or "").strip() for column in yushengtang_columns):
                device_count += 1
            if key not in key_to_detail_row:
                key_to_detail_row[key] = row_idx + 2
            if key not in key_to_status or status_priority.get(status, 0) > status_priority.get(key_to_status[key], 0):
                key_to_status[key] = status
            key_to_device_count[key] = min(2, max(key_to_device_count.get(key, 0), device_count))
            key_to_duplicate_flag[key] = key_to_duplicate_flag.get(key, False) or bool(str(row.get("疑似重复数值", "")).strip())

        row_labels = [f"{name}-{slot}" for name in cohort.names for slot in SLOT_LABELS]
        for row_idx, row_label in enumerate(row_labels, start=1):
            matrix_sheet.write(row_idx, 0, row_label, row_label_fmt)
            user_name, slot_label = row_label.rsplit("-", 1)
            for col_idx, date_label in enumerate(matrix_dates, start=1):
                key = (user_name, slot_label, date_label)
                if key not in key_to_status:
                    matrix_sheet.write_blank(row_idx, col_idx, None, blank_fmt)
                    continue
                status = key_to_status[key]
                detail_row = key_to_detail_row[key]
                if key_to_duplicate_flag.get(key, False):
                    cell_format = red_link_fmt
                else:
                    cell_format = {
                        "complete": green_link_fmt,
                        "incomplete": yellow_link_fmt,
                        "invalid": orange_link_fmt,
                        "suspicious": red_link_fmt,
                    }.get(status, blank_fmt)
                matrix_sheet.write_url(
                    row_idx,
                    col_idx,
                    f"internal:'详细记录'!A{detail_row}",
                    cell_format,
                    string=str(key_to_device_count.get(key, 1)),
                )

        matrix_sheet.freeze_panes(1, 1)
        detail_sheet.freeze_panes(1, 0)
        matrix_sheet.set_column(0, 0, 18)
        matrix_sheet.set_column(1, len(matrix_dates), 4.5)
        detail_sheet.set_column(0, 0, 12)
        detail_sheet.set_column(1, 1, 12)
        detail_sheet.set_column(2, 2, 10)
        detail_sheet.set_column(3, 5, 16)
        detail_sheet.set_column(6, len(detail_columns) - 1, 12)


def build_heatmap_matrix(slot_df: pd.DataFrame, cohort: CohortContext) -> pd.DataFrame:
    index_labels = [f"{name}-{slot}" for name in cohort.names for slot in SLOT_LABELS]
    matrix = pd.DataFrame(0, index=index_labels, columns=[day.strftime("%Y-%m-%d") for day in DATE_INDEX], dtype=int)
    if slot_df.empty:
        return matrix
    status_to_value = {"missing": 0, "incomplete": 1, "complete": 2, "suspicious": 3, "invalid": 1}
    for _, row in slot_df.iterrows():
        if row["slot_label"] not in SLOT_LABELS:
            continue
        label = f"{row['user_name']}-{row['slot_label']}"
        date_label = pd.Timestamp(row["date"]).strftime("%Y-%m-%d")
        value = status_to_value[row["slot_status"]]
        if bool(row.get("duplicate_numeric_flag", False)):
            value = 3
        matrix.loc[label, date_label] = max(matrix.loc[label, date_label], value)
    return matrix


def plot_heatmap(matrix: pd.DataFrame, output_path: Path, anomaly_points: pd.DataFrame | None = None) -> None:
    cmap = ListedColormap(["#f5f5f5", "#f6c85f", "#2ca02c", "#d62728"])
    norm = BoundaryNorm([-0.5, 0.5, 1.5, 2.5, 3.5], cmap.N)
    fig, ax = plt.subplots(figsize=(28, 18))
    im = ax.imshow(matrix.values, aspect="auto", cmap=cmap, norm=norm)
    ax.set_xticks(range(len(matrix.columns)))
    ax.set_xticklabels(matrix.columns, rotation=90, fontsize=6)
    slots_per_user = len(SLOT_LABELS)
    user_names = [str(label).rsplit("-", 1)[0] for label in matrix.index[::slots_per_user]]
    ytick_positions = [idx * slots_per_user + (slots_per_user - 1) / 2 for idx in range(len(user_names))]
    ax.set_yticks(ytick_positions)
    ax.set_yticklabels(user_names, fontsize=11)
    ax.tick_params(axis="y", length=0, pad=6)
    for boundary in range(slots_per_user, len(matrix.index), slots_per_user):
        ax.axhline(boundary - 0.5, color="#d9d9d9", linewidth=0.6)
    if anomaly_points is not None and not anomaly_points.empty:
        user_to_y = {user_name: ytick_positions[idx] for idx, user_name in enumerate(user_names)}
        date_to_x = {date_label: idx for idx, date_label in enumerate(matrix.columns)}
        xs: list[int] = []
        ys: list[float] = []
        for _, row in anomaly_points.iterrows():
            user_name = str(row["user_name"])
            date_label = pd.Timestamp(row["date"]).strftime("%Y-%m-%d")
            if user_name in user_to_y and date_label in date_to_x:
                xs.append(date_to_x[date_label])
                ys.append(user_to_y[user_name])
        if xs:
            ax.scatter(
                xs,
                ys,
                s=58,
                facecolors="none",
                edgecolors="#d62728",
                linewidths=1.5,
                marker="o",
                zorder=4,
            )
    ax.set_xlabel("日期")
    ax.set_ylabel("用户", fontsize=12)
    ax.set_title("2026Q1 打卡时段热力图（圆圈=疑似作弊, 0=缺失, 1=不完整, 2=有效, 3=异常）")
    cbar = fig.colorbar(im, ax=ax, ticks=[0, 1, 2, 3], fraction=0.015, pad=0.01)
    cbar.ax.set_yticklabels(["缺失", "不完整", "有效", "异常"])
    fig.tight_layout()
    fig.subplots_adjust(left=0.09)
    fig.savefig(output_path, dpi=200)
    plt.close(fig)


def plot_valid_day_bar(summary_df: pd.DataFrame, output_path: Path) -> None:
    ordered = summary_df.sort_values(["full_valid_days", "partial_valid_days"], ascending=[False, False])
    fig, ax = plt.subplots(figsize=(18, 8))
    positions = range(len(ordered))
    ax.bar(positions, ordered["full_valid_days"], label="完整有效天数", color="#2ca02c")
    ax.bar(
        positions,
        ordered["partial_valid_days"],
        bottom=ordered["full_valid_days"],
        label="部分有效天数",
        color="#f6c85f",
    )
    ax.set_xticks(list(positions))
    ax.set_xticklabels(ordered["user_name"], rotation=90)
    ax.set_ylabel("天数")
    ax.set_title("2026Q1 用户有效天数分布")
    ax.legend()
    fig.tight_layout()
    fig.savefig(output_path, dpi=200)
    plt.close(fig)


def write_excel_output(
    cohort: CohortContext,
    visits: pd.DataFrame,
    modality_rules: pd.DataFrame,
    sessions: pd.DataFrame,
    gap_summary: pd.DataFrame,
    gap_details: pd.DataFrame,
    day_slots: pd.DataFrame,
    day_summary: pd.DataFrame,
    user_summary: pd.DataFrame,
    source_summary: pd.DataFrame,
    anomalies: pd.DataFrame,
    file_scope: pd.DataFrame,
    heatmap_matrix: pd.DataFrame,
) -> None:
    with pd.ExcelWriter(WORKBOOK_OUTPUT, engine="xlsxwriter", datetime_format="yyyy-mm-dd hh:mm:ss") as writer:
        pd.DataFrame({"user_name": cohort.names}).to_excel(writer, sheet_name="cohort", index=False)
        visits.sort_values(["user_name", "collected_at", "source_vendor"]).to_excel(writer, sheet_name="visits", index=False)
        modality_rules.to_excel(writer, sheet_name="modality_rules", index=False)
        sessions.sort_values(["user_name", "cluster_start"]).to_excel(writer, sheet_name="sessions", index=False)
        gap_summary.to_excel(writer, sheet_name="gap_summary", index=False)
        gap_details.sort_values(["user_name", "date", "gap_minutes"]).to_excel(writer, sheet_name="gap_details", index=False)
        day_slots.sort_values(["user_name", "date", "cluster_start"]).to_excel(writer, sheet_name="day_slots", index=False)
        day_summary.sort_values(["user_name", "date"]).to_excel(writer, sheet_name="day_summary", index=False)
        user_summary.sort_values(["full_valid_days", "partial_valid_days"], ascending=[False, False]).to_excel(writer, sheet_name="user_summary", index=False)
        source_summary.to_excel(writer, sheet_name="source_summary", index=False)
        anomalies.to_excel(writer, sheet_name="anomalies", index=False)
        file_scope.sort_values(["source_vendor", "source_path"]).to_excel(writer, sheet_name="file_scope", index=False)
        heatmap_matrix.reset_index().rename(columns={"index": "user_slot"}).to_excel(
            writer, sheet_name="heatmap_matrix", index=False
        )


def frame_to_markdown(df: pd.DataFrame, index: bool = False) -> str:
    if df.empty:
        return "(空)"
    render_df = df.copy()
    if index:
        render_df = render_df.reset_index()
    render_df = render_df.fillna("")
    columns = [str(column) for column in render_df.columns]
    rows = [
        ["" if value is None else str(value) for value in row]
        for row in render_df.astype(object).itertuples(index=False, name=None)
    ]
    header = "| " + " | ".join(columns) + " |"
    separator = "| " + " | ".join(["---"] * len(columns)) + " |"
    body = ["| " + " | ".join(row) + " |" for row in rows]
    return "\n".join([header, separator, *body])


def build_generated_report(
    cohort: CohortContext,
    modality_rules: pd.DataFrame,
    gap_summary: pd.DataFrame,
    source_summary: pd.DataFrame,
    user_summary: pd.DataFrame,
    anomalies: pd.DataFrame,
    file_scope: pd.DataFrame,
) -> str:
    total_days = len(DATE_INDEX)
    zhongke_required = ",".join(
        modality_rules.loc[(modality_rules["source_vendor"] == "zhongke") & (modality_rules["is_required"]), "modality"].tolist()
    )
    yst_required = ",".join(
        modality_rules.loc[(modality_rules["source_vendor"] == "yushengtang") & (modality_rules["is_required"]), "modality"].tolist()
    )
    gap_lookup = dict(zip(gap_summary["metric"], gap_summary["value"])) if not gap_summary.empty else {}
    top_missing = user_summary.sort_values(["missing_days", "partial_valid_days"], ascending=[False, True]).head(10)
    scope_preview = file_scope.groupby("source_vendor").agg(
        matched_files=("source_path", "nunique"),
        earliest=("min_collected_at", "min"),
        latest=("max_collected_at", "max"),
    )
    lines = [
        "# 2026Q1 数据质量分析（脚本自动生成）",
        "",
        f"- cohort 用户数：{len(cohort.names)}",
        f"- 分析日期范围：{Q1_START.date()} ~ {(Q2_START - pd.Timedelta(days=1)).date()}（共 {total_days} 天）",
        "",
        "## 1. 文件范围",
        "",
        frame_to_markdown(scope_preview, index=True),
        "",
        "## 2. 规则",
        "",
        f"- 中科要求模态：{zhongke_required or '(无)'}",
        f"- 玉生堂要求模态：{yst_required or '(无)'}",
        f"- 10 分钟内 3 次及以上聚集记录视为异常。",
        f"- 固定间隔阈值：{gap_lookup.get('gap_threshold_minutes', 0)} 分钟。",
        f"- 典型间隔：P25={gap_lookup.get('p25_gap_minutes', 0)} 分钟, 中位数={gap_lookup.get('median_gap_minutes', 0)} 分钟, P75={gap_lookup.get('p75_gap_minutes', 0)} 分钟。",
        "",
        "## 3. 概览",
        "",
        frame_to_markdown(source_summary, index=False) if not source_summary.empty else "(无源级汇总)",
        "",
        "## 4. 缺失最高的 10 位用户",
        "",
        frame_to_markdown(
            top_missing[
                ["user_name", "full_valid_days", "partial_valid_days", "missing_days", "triplicate_clusters", "spacing_violations"]
            ],
            index=False,
        ),
        "",
        "## 5. 异常统计",
        "",
        f"- 异常 cluster 数：{len(anomalies)}",
        "",
    ]
    return "\n".join(lines)


def main() -> None:
    ensure_output_dirs()
    cohort = build_cohort_context()

    zhongke_visits, zhongke_file_scope = parse_zhongke_records(cohort)
    yushengtang_visits, yushengtang_file_scope = parse_yushengtang_records(cohort)
    visits = pd.concat([zhongke_visits, yushengtang_visits], ignore_index=True, sort=False)
    if visits.empty:
        raise RuntimeError("Q1 2026 未解析到任何 cohort 记录。")

    for column in ["ask", "pulse", "tongue", "face", "voice", "basic", "western", "number_order", "pdf", "has_case_dir"]:
        if column not in visits.columns:
            visits[column] = False
        visits[column] = visits[column].fillna(False).astype(bool)

    visits["collected_at"] = pd.to_datetime(visits["collected_at"])
    visits = visits.sort_values(["user_name", "collected_at", "source_vendor"]).reset_index(drop=True)
    modality_rules = determine_required_modalities(visits)
    visits = apply_visit_quality_flags(visits, modality_rules)
    visits = detect_duplicate_numeric_visits(visits)
    sessions, anomalies = cluster_sessions(visits)
    gap_threshold_minutes, gap_details = compute_gap_threshold_minutes(sessions)
    gap_summary = build_gap_summary(gap_details, gap_threshold_minutes)
    day_slots, day_summary = assign_day_slots(sessions, gap_threshold_minutes)
    user_summary = build_user_summary(day_summary, cohort)
    source_summary = build_source_summary(visits, modality_rules)
    file_scope = pd.concat([zhongke_file_scope, yushengtang_file_scope], ignore_index=True, sort=False).fillna("")
    heatmap_matrix = build_heatmap_matrix(day_slots, cohort)
    detail_export_df = build_detail_export(day_slots, sessions, visits)

    plot_heatmap(heatmap_matrix, PLOT_DIR / "q1_2026_slot_heatmap.png", anomalies)
    plot_valid_day_bar(user_summary, PLOT_DIR / "q1_2026_valid_day_bar.png")
    write_excel_output(
        cohort=cohort,
        visits=visits,
        modality_rules=modality_rules,
        sessions=sessions,
        gap_summary=gap_summary,
        gap_details=gap_details,
        day_slots=day_slots,
        day_summary=day_summary,
        user_summary=user_summary,
        source_summary=source_summary,
        anomalies=anomalies,
        file_scope=file_scope,
        heatmap_matrix=heatmap_matrix,
    )
    write_linked_export(cohort, day_slots, detail_export_df)

    generated_report = build_generated_report(
        cohort=cohort,
        modality_rules=modality_rules,
        gap_summary=gap_summary,
        source_summary=source_summary,
        user_summary=user_summary,
        anomalies=anomalies,
        file_scope=file_scope,
    )
    GENERATED_REPORT_OUTPUT.write_text(generated_report, encoding="utf-8")

    print(f"Saved workbook: {WORKBOOK_OUTPUT}")
    print(f"Saved linked export: {LINKED_EXPORT_OUTPUT}")
    print(f"Saved heatmap: {PLOT_DIR / 'q1_2026_slot_heatmap.png'}")
    print(f"Saved bar chart: {PLOT_DIR / 'q1_2026_valid_day_bar.png'}")
    print(f"Saved generated report: {GENERATED_REPORT_OUTPUT}")


if __name__ == "__main__":
    main()
