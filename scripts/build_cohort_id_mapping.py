from __future__ import annotations

import json
import math
import re
from collections import Counter, defaultdict
from io import BytesIO
from pathlib import Path

import openpyxl
import pandas as pd
import xlrd


def u(text: str) -> str:
    return text.encode("ascii").decode("unicode_escape")


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_ROOT = PROJECT_ROOT / u("\\u56db\\u8bca\\u4eea\\u6570\\u636e\\u6574\\u7406")
ZHONGKE_REFERENCE_DIR = DATA_ROOT / u("\\u4e2d\\u79d1\\u56db\\u8bca\\u4eea") / "2025.11.09-2025.12.10"
YUSHENGTANG_ROOT = DATA_ROOT / u("\\u7389\\u751f\\u5802\\u56db\\u8bca\\u4eea")
OUTPUT_DIR = PROJECT_ROOT / "datasets" / "id_mapping"
OUTPUT_CSV = OUTPUT_DIR / "cohort_20_name_to_ids.csv"
OUTPUT_MD = OUTPUT_DIR / "cohort_20_name_to_ids.md"
EXISTING_ZHONGKE_ROWS = PROJECT_ROOT / "datasets" / "caid_name_check" / "zhongke_name_mobile_caseid_rows.csv"

SHEET_PULSE = u("\\u8109\\u8bca\\u6570\\u636e")
SHEET_TONGUE = u("\\u820c\\u8bca\\u6570\\u636e")
SHEET_FACE = u("\\u9762\\u8bca\\u6570\\u636e")
SHEET_ASK = u("\\u95ee\\u8bca\\u6c47\\u603b")

ALIAS_TO_CANONICAL = {
    "aln": "爱丽娜",
    "艾琳娜aln": "爱丽娜",
    "艾琳娜": "爱丽娜",
    "zyn": "张雨楠",
    "ZYN": "张雨楠",
    "lyl": "李昱霖",
    "蒋": "蒋广祥",
}


def normalize_name(value: object) -> str:
    text = str(value or "").strip()
    text = text.replace("\n", "").replace("\r", "")
    text = re.sub(r"\s+", "", text)
    return ALIAS_TO_CANONICAL.get(text, text)


def normalize_phone(value: object) -> str:
    raw = str(value or "")
    matches = re.findall(r"1\d{10}", raw)
    if matches:
        uniq = []
        for match in matches:
            if match not in uniq:
                uniq.append(match)
        if len(uniq) == 1:
            return uniq[0]
        return "|".join(uniq)
    digits = re.sub(r"\D+", "", raw)
    if not digits:
        return ""
    if len(digits) == 10 and digits.startswith("1"):
        return digits
    if len(digits) > 11 and digits.startswith("86"):
        digits = digits[-11:]
    if len(digits) > 11 and digits.startswith("0"):
        digits = digits[-11:]
    return digits


def coerce_id(value: object) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def parse_datetime_value(value: object) -> pd.Timestamp | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        numeric = float(value)
        if math.isnan(numeric):
            return None
        if 20000 <= numeric <= 80000:
            try:
                return pd.Timestamp(pd.to_datetime(numeric, unit="D", origin="1899-12-30"))
            except Exception:
                pass
    try:
        ts = pd.to_datetime(value)
    except Exception:
        return None
    if pd.isna(ts):
        return None
    return pd.Timestamp(ts)


def read_text_relaxed(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="ignore")


def load_json_relaxed(path: Path) -> object | None:
    if not path.exists():
        return None
    try:
        return json.loads(read_text_relaxed(path))
    except Exception:
        return None


def candidate_name_columns(headers: list[str]) -> list[int]:
    indices: list[int] = []
    for idx, header in enumerate(headers):
        text = normalize_name(header)
        if text in {u("\\u59d3\\u540d"), "CaName"} or "姓名" in text:
            indices.append(idx)
    return indices


def candidate_phone_columns(headers: list[str]) -> list[int]:
    indices: list[int] = []
    for idx, header in enumerate(headers):
        text = normalize_name(header)
        if text in {u("\\u624b\\u673a\\u53f7"), "CaMobile"} or "手机" in text:
            indices.append(idx)
    return indices


def candidate_caseid_columns(headers: list[str]) -> list[int]:
    indices: list[int] = []
    for idx, header in enumerate(headers):
        text = normalize_name(header)
        if text in {u("\\u75c5\\u4f8b\\u53f7"), u("\\u75c5\\u4f8bID"), u("\\u75c5\\u5386\\u7f16\\u53f7")}:
            indices.append(idx)
    return indices


def iter_zhongke_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    excel_paths = sorted(path for path in ZHONGKE_REFERENCE_DIR.rglob("*") if path.is_file() and path.suffix.lower() in {".xls", ".xlsx"})
    for path in excel_paths:
        raw = path.read_bytes()
        head = raw[:8]
        try:
            if head.startswith(b"PK"):
                workbook = openpyxl.load_workbook(BytesIO(raw), read_only=True, data_only=True)
                try:
                    for worksheet in workbook.worksheets:
                        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                            values = list(row)
                            if not values:
                                continue
                            headers = [normalize_name(value) for value in values]
                            name_cols = candidate_name_columns(headers)
                            phone_cols = candidate_phone_columns(headers)
                            caseid_cols = candidate_caseid_columns(headers)
                            if name_cols and phone_cols and caseid_cols:
                                for data_row in worksheet.iter_rows(min_row=row_idx + 1, values_only=True):
                                    cells = list(data_row)
                                    if len(cells) <= max(name_cols + phone_cols + caseid_cols):
                                        continue
                                    name = normalize_name(cells[name_cols[0]])
                                    phone = normalize_phone(cells[phone_cols[0]])
                                    case_id = coerce_id(cells[caseid_cols[0]])
                                    collected_at = None
                                    if len(cells) > 6:
                                        collected_at = parse_datetime_value(cells[6])
                                    if not name or not case_id:
                                        continue
                                    rows.append(
                                        {
                                            "name": name,
                                            "phone": phone,
                                            "case_id": case_id,
                                            "collected_at": "" if collected_at is None else collected_at.isoformat(),
                                            "source_file": str(path),
                                            "sheet_name": worksheet.title,
                                        }
                                    )
                                break
                finally:
                    workbook.close()
            elif head == bytes.fromhex("D0CF11E0A1B11AE1"):
                workbook = xlrd.open_workbook(file_contents=raw, on_demand=True)
                try:
                    for sheet_name in workbook.sheet_names():
                        sheet = workbook.sheet_by_name(sheet_name)
                        header_row_idx = None
                        name_cols: list[int] = []
                        phone_cols: list[int] = []
                        caseid_cols: list[int] = []
                        for row_idx in range(min(sheet.nrows, 12)):
                            headers = [normalize_name(value) for value in sheet.row_values(row_idx)]
                            name_cols = candidate_name_columns(headers)
                            phone_cols = candidate_phone_columns(headers)
                            caseid_cols = candidate_caseid_columns(headers)
                            if name_cols and phone_cols and caseid_cols:
                                header_row_idx = row_idx
                                break
                        if header_row_idx is None:
                            continue
                        for row_idx in range(header_row_idx + 1, sheet.nrows):
                            cells = sheet.row_values(row_idx)
                            if len(cells) <= max(name_cols + phone_cols + caseid_cols):
                                continue
                            name = normalize_name(cells[name_cols[0]])
                            phone = normalize_phone(cells[phone_cols[0]])
                            case_id = coerce_id(cells[caseid_cols[0]])
                            collected_at = None
                            if len(cells) > 6:
                                collected_at = parse_datetime_value(cells[6])
                            if not name or not case_id:
                                continue
                            rows.append(
                                {
                                    "name": name,
                                    "phone": phone,
                                    "case_id": case_id,
                                    "collected_at": "" if collected_at is None else collected_at.isoformat(),
                                    "source_file": str(path),
                                    "sheet_name": sheet_name,
                                }
                            )
                finally:
                    workbook.release_resources()
        except Exception:
            continue
    dedup = pd.DataFrame(rows).drop_duplicates()
    return dedup.to_dict("records")


def parse_yushengtang_records() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for archive_path in sorted(YUSHENGTANG_ROOT.rglob("customerArchive.json")):
        payload = load_json_relaxed(archive_path)
        if not isinstance(payload, dict):
            continue
        case_dir = archive_path.parent
        treat_number = coerce_id(payload.get("TreatNumber") or case_dir.name)
        caid = coerce_id(payload.get("CaId"))
        name = normalize_name(payload.get("CaName"))
        phone = normalize_phone(payload.get("CaMobile"))
        if not name and not phone and not caid:
            continue
        rows.append(
            {
                "name": name,
                "phone": phone,
                "caid": caid,
                "treat_number": treat_number,
                "source_file": str(archive_path),
            }
        )
    dedup = pd.DataFrame(rows).drop_duplicates()
    return dedup.to_dict("records")


def build_zhongke_summary(records: list[dict[str, str]], roster: list[str]) -> pd.DataFrame:
    df = pd.DataFrame(records)
    if df.empty:
        return pd.DataFrame(columns=["roster_name", "zhongke_table_name", "zhongke_phone", "zhongke_case_id_count", "zhongke_first_case_id", "evidence"])

    rows: list[dict[str, str]] = []
    roster_set = set(roster)
    direct = df[df["name"].isin(roster_set)].copy()
    for name in roster:
        subset = direct[direct["name"] == name]
        if subset.empty:
            rows.append(
                {
                    "roster_name": name,
                    "zhongke_table_name": "",
                    "zhongke_phone": "",
                    "zhongke_case_id_count": 0,
                    "zhongke_first_case_id": "",
                    "evidence": "",
                }
            )
            continue
        phones = [phone for phone in sorted(subset["phone"].dropna().astype(str).unique()) if phone]
        rows.append(
            {
                "roster_name": name,
                "zhongke_table_name": name,
                "zhongke_phone": "|".join(phones),
                "zhongke_case_id_count": int(subset["case_id"].nunique()),
                "zhongke_first_case_id": str(subset["case_id"].dropna().astype(str).iloc[0]),
                "evidence": "direct_name_in_table",
            }
        )

    summary = pd.DataFrame(rows)

    unresolved = summary[summary["zhongke_phone"] == ""]["roster_name"].tolist()
    if unresolved and EXISTING_ZHONGKE_ROWS.exists():
        all_rows = pd.read_csv(EXISTING_ZHONGKE_ROWS, dtype=str).fillna("")
        for missing_name in unresolved:
            subset = all_rows[all_rows["name"].astype(str).str.replace(r"\s+", "", regex=True) == missing_name]
            if subset.empty:
                continue
            phones = [phone for phone in sorted({normalize_phone(phone) for phone in subset["mobile"].astype(str)}) if phone]
            summary.loc[summary["roster_name"] == missing_name, "zhongke_table_name"] = missing_name
            summary.loc[summary["roster_name"] == missing_name, "zhongke_phone"] = "|".join(phones)
            summary.loc[summary["roster_name"] == missing_name, "zhongke_case_id_count"] = int(subset["case_id"].nunique())
            summary.loc[summary["roster_name"] == missing_name, "zhongke_first_case_id"] = str(subset["case_id"].iloc[0])
            summary.loc[summary["roster_name"] == missing_name, "evidence"] = "direct_name_in_other_zhongke_tables"

    unresolved = summary[summary["zhongke_phone"] == ""]["roster_name"].tolist()
    if unresolved:
        alias_df = df[~df["name"].isin(roster_set)].copy()
        alias_phone_map = (
            alias_df.groupby(["name", "phone"], dropna=False)["case_id"]
            .nunique()
            .reset_index(name="case_count")
            .sort_values(["name", "case_count"], ascending=[True, False])
        )
        known_phone_to_name = {}
        for _, row in summary.iterrows():
            for phone in str(row["zhongke_phone"]).split("|"):
                if phone:
                    known_phone_to_name[phone] = row["roster_name"]

        yst_phone_map = build_yushengtang_phone_map()
        for missing_name in unresolved:
            yst_phone = yst_phone_map.get(missing_name, "")
            if not yst_phone:
                continue
            candidate = alias_phone_map[alias_phone_map["phone"] == yst_phone]
            if candidate.empty:
                continue
            best = candidate.iloc[0]
            summary.loc[summary["roster_name"] == missing_name, "zhongke_table_name"] = best["name"]
            summary.loc[summary["roster_name"] == missing_name, "zhongke_phone"] = yst_phone
            summary.loc[summary["roster_name"] == missing_name, "zhongke_case_id_count"] = int(df[(df["name"] == best["name"]) & (df["phone"] == yst_phone)]["case_id"].nunique())
            first_case = df[(df["name"] == best["name"]) & (df["phone"] == yst_phone)]["case_id"].astype(str).iloc[0]
            summary.loc[summary["roster_name"] == missing_name, "zhongke_first_case_id"] = first_case
            summary.loc[summary["roster_name"] == missing_name, "evidence"] = "matched_by_phone_to_yushengtang"
    return summary


_YUSHENGTANG_PHONE_CACHE: dict[str, str] | None = None


def build_yushengtang_phone_map() -> dict[str, str]:
    global _YUSHENGTANG_PHONE_CACHE
    if _YUSHENGTANG_PHONE_CACHE is not None:
        return _YUSHENGTANG_PHONE_CACHE
    records = parse_yushengtang_records()
    df = pd.DataFrame(records)
    mapping: dict[str, str] = {}
    if not df.empty:
        for name, subset in df.groupby("name"):
            phones = [phone for phone in subset["phone"].astype(str).unique() if phone]
            if len(phones) == 1:
                mapping[name] = phones[0]
    _YUSHENGTANG_PHONE_CACHE = mapping
    return mapping


def build_yushengtang_summary(roster: list[str]) -> pd.DataFrame:
    records = parse_yushengtang_records()
    df = pd.DataFrame(records)
    rows: list[dict[str, str]] = []
    if df.empty:
        return pd.DataFrame(columns=["roster_name", "yst_name", "yst_phone", "yst_caid", "yst_treat_count"])

    for name in roster:
        subset = df[df["name"] == name].copy()
        if subset.empty:
            rows.append(
                {
                    "roster_name": name,
                    "yst_name": "",
                    "yst_phone": "",
                    "yst_caid": "",
                    "yst_treat_count": 0,
                }
            )
            continue
        phones = [phone for phone in sorted(subset["phone"].dropna().astype(str).unique()) if phone]
        caids = [caid for caid in sorted(subset["caid"].dropna().astype(str).unique()) if caid]
        rows.append(
            {
                "roster_name": name,
                "yst_name": name,
                "yst_phone": "|".join(phones),
                "yst_caid": "|".join(caids),
                "yst_treat_count": int(subset["treat_number"].nunique()),
            }
        )
    return pd.DataFrame(rows)


def build_final_mapping(roster: list[str]) -> pd.DataFrame:
    zhongke_records = iter_zhongke_rows()
    zhongke = build_zhongke_summary(zhongke_records, roster)
    yst = build_yushengtang_summary(roster)
    final = zhongke.merge(yst, on="roster_name", how="left")
    final["zhongke_user_id_rule"] = final.apply(
        lambda row: f"{row['roster_name']}+{row['zhongke_phone']}" if row["zhongke_phone"] else "",
        axis=1,
    )
    final["yst_user_id_rule"] = final.apply(
        lambda row: f"{row['roster_name']}+{row['yst_phone']}" if row["yst_phone"] else "",
        axis=1,
    )
    final["zhongke_visit_id_rule"] = u("\\u75c5\\u4f8b\\u53f7")
    final["yst_visit_id_rule"] = "TreatNumber"
    final["notes"] = final.apply(make_note, axis=1)
    return final.sort_values("roster_name").reset_index(drop=True)


def make_note(row: pd.Series) -> str:
    notes: list[str] = []
    if row["evidence"] == "direct_name_in_table":
        notes.append("中科姓名直接见于表格")
    elif row["evidence"] == "direct_name_in_other_zhongke_tables":
        notes.append("中科姓名未在参考目录表内直接抽出，但在其他中科表内可直接确认")
    elif row["evidence"] == "matched_by_phone_to_yushengtang":
        notes.append(f"中科表内别名 {row['zhongke_table_name']} 通过手机号与玉生堂对应")
    else:
        notes.append("中科未在表内直接确认")

    if not row["yst_phone"]:
        notes.append("玉生堂未找到姓名+手机号记录")
    if row["yst_phone"] and row["zhongke_phone"] and row["yst_phone"] != row["zhongke_phone"]:
        notes.append("两侧手机号不一致")
    return "；".join(notes)


def write_outputs(df: pd.DataFrame) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")

    header = "| " + " | ".join(df.columns) + " |"
    separator = "| " + " | ".join(["---"] * len(df.columns)) + " |"
    body = [
        "| " + " | ".join(str(value).replace("\n", " ") for value in row) + " |"
        for row in df.itertuples(index=False, name=None)
    ]
    lines = [
        "# 20人中科/玉生堂ID对照表",
        "",
        f"- 中科用户ID规则：`姓名 + 手机号`；visit级ID：`病例号`",
        f"- 玉生堂用户ID规则：`姓名 + 手机号`；visit级ID：`TreatNumber`，`CaId` 作为辅助来源ID保留",
        "",
        header,
        separator,
        *body,
        "",
    ]
    OUTPUT_MD.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    roster = sorted(path.name for path in ZHONGKE_REFERENCE_DIR.iterdir() if path.is_dir())
    final = build_final_mapping(roster)
    write_outputs(final)
    print(final.to_string(index=False))
    print(f"\nWrote: {OUTPUT_CSV}")
    print(f"Wrote: {OUTPUT_MD}")


if __name__ == "__main__":
    main()
