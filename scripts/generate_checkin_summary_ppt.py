from __future__ import annotations

import html
import zipfile
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from PIL import Image


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_ROOT / "datasets" / "organized_checkin_matrix"
WORKBOOK = DATA_DIR / "cohort_checkin_matrix_20251108.xlsx"
HEATMAP = DATA_DIR / "cohort_checkin_heatmap_20251108.png"
OUTPUT = DATA_DIR / "cohort_checkin_summary_20251108.pptx"

RULE_VERSION = "cohort_rule_v1_20260422"

SLIDE_W = 12192000
SLIDE_H = 6858000


def emu(inches: float) -> int:
    return int(inches * 914400)


def esc(text: object) -> str:
    return html.escape("" if text is None else str(text))


def date_text(value: object, epoch) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        return from_excel(value, epoch).strftime("%Y-%m-%d")
    return str(value)


def load_stats() -> dict[str, object]:
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb["\u8be6\u7ec6\u8bb0\u5f55"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    users = set()
    dates = []
    status = Counter()
    slots = Counter()
    mismatch = 0
    duplicate_numeric = 0
    remark_heads = Counter()

    for row in ws.iter_rows(min_row=2, values_only=True):
        users.add(row[idx["\u59d3\u540d"]])
        dates.append(row[idx["\u65e5\u671f"]])
        status[row[idx["\u72b6\u6001"]]] += 1
        slots[row[idx["\u65f6\u6bb5"]]] += 1
        if row[idx["\u59d3\u540d\u4e0d\u4e00\u81f4"]]:
            mismatch += 1
        if row[idx["\u7591\u4f3c\u91cd\u590d\u6570\u503c"]]:
            duplicate_numeric += 1
        remark = row[idx["\u5907\u6ce8"]]
        if remark:
            remark_heads[str(remark).split("\uff1b")[0]] += 1

    date_strings = [date_text(v, wb.epoch) for v in dates if v is not None]
    matrix = wb["\u6253\u5361\u77e9\u9635"]
    complete = status.get("complete", 0)
    incomplete = status.get("incomplete", 0)
    suspicious = status.get("suspicious", 0)
    invalid = status.get("invalid", 0)
    total = complete + incomplete + suspicious + invalid
    return {
        "users": len(users),
        "date_min": min(date_strings),
        "date_max": max(date_strings),
        "active_dates": len(set(date_strings)),
        "status": status,
        "slots": slots,
        "mismatch": mismatch,
        "duplicate_numeric": duplicate_numeric,
        "total": total,
        "complete_rate": complete / total if total else 0,
        "matrix_rows": matrix.max_row,
        "matrix_cols": matrix.max_column,
        "top_remarks": remark_heads.most_common(5),
    }


def tx_body(text: str, font_size: int = 24, bold: bool = False, color: str = "333333") -> str:
    lines = str(text).splitlines() or [""]
    paras = []
    for line in lines:
        paras.append(
            f"""
            <a:p>
              <a:r>
                <a:rPr lang="zh-CN" sz="{font_size * 100}" b="{1 if bold else 0}">
                  <a:solidFill><a:srgbClr val="{color}"/></a:solidFill>
                  <a:latin typeface="Microsoft YaHei"/>
                  <a:ea typeface="Microsoft YaHei"/>
                </a:rPr>
                <a:t>{esc(line)}</a:t>
              </a:r>
              <a:endParaRPr lang="zh-CN" sz="{font_size * 100}"/>
            </a:p>
            """
        )
    return "<p:txBody><a:bodyPr wrap=\"square\"/><a:lstStyle/>" + "".join(paras) + "</p:txBody>"


def text_box(shape_id: int, x: int, y: int, cx: int, cy: int, text: str, size: int = 24, bold: bool = False, color: str = "333333") -> str:
    return f"""
    <p:sp>
      <p:nvSpPr><p:cNvPr id="{shape_id}" name="Text {shape_id}"/><p:cNvSpPr txBox="1"/><p:nvPr/></p:nvSpPr>
      <p:spPr><a:xfrm><a:off x="{x}" y="{y}"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm><a:prstGeom prst="rect"><a:avLst/></a:prstGeom><a:noFill/></p:spPr>
      {tx_body(text, size, bold, color)}
    </p:sp>
    """


def rect(shape_id: int, x: int, y: int, cx: int, cy: int, fill: str, line: str = "FFFFFF") -> str:
    return f"""
    <p:sp>
      <p:nvSpPr><p:cNvPr id="{shape_id}" name="Rect {shape_id}"/><p:cNvSpPr/><p:nvPr/></p:nvSpPr>
      <p:spPr>
        <a:xfrm><a:off x="{x}" y="{y}"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm>
        <a:prstGeom prst="roundRect"><a:avLst/></a:prstGeom>
        <a:solidFill><a:srgbClr val="{fill}"/></a:solidFill>
        <a:ln><a:solidFill><a:srgbClr val="{line}"/></a:solidFill></a:ln>
      </p:spPr>
    </p:sp>
    """


def image_xml(shape_id: int, rel_id: str, x: int, y: int, cx: int, cy: int) -> str:
    return f"""
    <p:pic>
      <p:nvPicPr><p:cNvPr id="{shape_id}" name="matrix.png"/><p:cNvPicPr/><p:nvPr/></p:nvPicPr>
      <p:blipFill><a:blip r:embed="{rel_id}"/><a:stretch><a:fillRect/></a:stretch></p:blipFill>
      <p:spPr><a:xfrm><a:off x="{x}" y="{y}"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm><a:prstGeom prst="rect"><a:avLst/></a:prstGeom></p:spPr>
    </p:pic>
    """


def slide_xml(shapes: list[str]) -> str:
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<p:sld xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
       xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
       xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main">
  <p:cSld><p:spTree>
    <p:nvGrpSpPr><p:cNvPr id="1" name=""/><p:cNvGrpSpPr/><p:nvPr/></p:nvGrpSpPr>
    <p:grpSpPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/><a:chOff x="0" y="0"/><a:chExt cx="0" cy="0"/></a:xfrm></p:grpSpPr>
    {''.join(shapes)}
  </p:spTree></p:cSld>
  <p:clrMapOvr><a:masterClrMapping/></p:clrMapOvr>
</p:sld>"""


def rels_xml(rels: list[tuple[str, str, str]]) -> str:
    body = "\n".join(
        f'<Relationship Id="{rid}" Type="{rtype}" Target="{target}"/>' for rid, rtype, target in rels
    )
    return f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">{body}</Relationships>'


def card(shape_id: int, x: float, y: float, w: float, h: float, title: str, body: str, color: str) -> list[str]:
    xi, yi, wi, hi = emu(x), emu(y), emu(w), emu(h)
    return [
        rect(shape_id, xi, yi, wi, hi, color, "D9E2EC"),
        text_box(shape_id + 100, xi + emu(0.18), yi + emu(0.12), wi - emu(0.36), emu(0.35), title, 15, True, "334155"),
        text_box(shape_id + 200, xi + emu(0.18), yi + emu(0.50), wi - emu(0.36), hi - emu(0.60), body, 25, True, "0F172A"),
    ]


def build_slides(stats: dict[str, object]) -> list[tuple[str, str | None]]:
    status = stats["status"]
    slots = stats["slots"]
    total = stats["total"]
    complete_rate = stats["complete_rate"]

    slide1 = slide_xml([
        rect(2, 0, 0, SLIDE_W, SLIDE_H, "F8FAFC", "F8FAFC"),
        text_box(3, emu(0.75), emu(0.75), emu(11.8), emu(0.9), "\u56db\u8bca\u4eea\u6253\u5361\u6570\u636e\u7edf\u8ba1\u6c47\u62a5", 34, True, "0F172A"),
        text_box(4, emu(0.78), emu(1.75), emu(11.2), emu(1.0), "2025-11-08 \u4ee5\u540e\uff0c20 \u4eba\u7eb5\u5411\u91c7\u96c6\u6570\u636e\n\u57fa\u4e8e\u5df2\u51bb\u7ed3\u89c4\u5219\u7248\u672c\uff1a" + RULE_VERSION, 22, False, "334155"),
        *card(10, 0.8, 3.2, 2.5, 1.2, "\u7528\u6237\u6570", str(stats["users"]), "E0F2FE"),
        *card(20, 3.55, 3.2, 2.5, 1.2, "\u6709\u8bb0\u5f55\u65e5\u671f", str(stats["active_dates"]), "DCFCE7"),
        *card(30, 6.3, 3.2, 2.5, 1.2, "\u8be6\u7ec6\u8bb0\u5f55", str(total), "FEF3C7"),
        *card(40, 9.05, 3.2, 2.5, 1.2, "\u5b8c\u6574\u7387", f"{complete_rate:.1%}", "FCE7F3"),
        text_box(50, emu(0.8), emu(5.25), emu(11.6), emu(0.8), f"\u65e5\u671f\u8303\u56f4\uff1a{stats['date_min']} \u81f3 {stats['date_max']}", 20, False, "475569"),
    ])

    status_lines = [
        f"complete\uff1a{status.get('complete', 0)}",
        f"incomplete\uff1a{status.get('incomplete', 0)}",
        f"suspicious\uff1a{status.get('suspicious', 0)}",
        f"invalid\uff1a{status.get('invalid', 0)}",
    ]
    slot_lines = [
        f"\u65e9\uff1a{slots.get('\u65e9', 0)}",
        f"\u4e2d\uff1a{slots.get('\u4e2d', 0)}",
        f"\u665a\uff1a{slots.get('\u665a', 0)}",
    ]
    top_remarks = "\n".join(f"{name}\uff1a{count}" for name, count in stats["top_remarks"])
    slide2 = slide_xml([
        rect(2, 0, 0, SLIDE_W, SLIDE_H, "FFFFFF", "FFFFFF"),
        text_box(3, emu(0.55), emu(0.35), emu(12.0), emu(0.5), "\u6838\u5fc3\u7edf\u8ba1", 28, True, "0F172A"),
        *card(10, 0.7, 1.25, 3.4, 2.3, "\u8d28\u91cf\u72b6\u6001", "\n".join(status_lines), "EFF6FF"),
        *card(20, 4.45, 1.25, 3.4, 2.3, "\u65e5\u5185\u65f6\u6bb5", "\n".join(slot_lines), "F0FDF4"),
        *card(30, 8.2, 1.25, 3.4, 2.3, "\u6807\u8bb0\u60c5\u51b5", f"\u59d3\u540d\u4e0d\u4e00\u81f4\uff1a{stats['mismatch']}\n\u7591\u4f3c\u91cd\u590d\u6570\u503c\uff1a{stats['duplicate_numeric']}", "FFF7ED"),
        text_box(40, emu(0.75), emu(4.2), emu(11.4), emu(1.45), "\u4e3b\u8981\u5f02\u5e38\u5907\u6ce8\n" + top_remarks, 20, False, "334155"),
    ])

    im = Image.open(HEATMAP)
    img_w, img_h = im.size
    max_w, max_h = emu(12.2), emu(5.35)
    scale = min(max_w / img_w, max_h / img_h)
    cx, cy = int(img_w * scale), int(img_h * scale)
    x = int((SLIDE_W - cx) / 2)
    y = emu(1.15)
    slide3 = slide_xml([
        rect(2, 0, 0, SLIDE_W, SLIDE_H, "FFFFFF", "FFFFFF"),
        text_box(3, emu(0.45), emu(0.3), emu(12.2), emu(0.45), "\u6253\u5361\u77e9\u9635\u56fe", 27, True, "0F172A"),
        image_xml(4, "rId2", x, y, cx, cy),
        text_box(5, emu(0.55), emu(6.55), emu(12.0), emu(0.35), "\u7eb5\u8f74\uff1a\u7528\u6237\uff1b\u6a2a\u8f74\uff1a\u65e5\u671f\uff1b\u989c\u8272\u53cd\u6620\u5b8c\u6574\u3001\u4e0d\u5b8c\u6574\u3001\u7591\u4f3c\u5f02\u5e38\u7b49\u72b6\u6001\u3002", 15, False, "475569"),
    ])

    rules = "\n".join([
        "1. \u7edf\u8ba1\u8303\u56f4\uff1a2025-11-08 \u4ee5\u540e\uff0c\u6307\u5b9a 20 \u4eba\u3002",
        "2. \u65e9/\u4e2d/\u665a\u4e0d\u91c7\u7528\u56fa\u5b9a\u65f6\u95f4\u7a97\uff0c\u800c\u662f\u5f53\u5929\u7b2c 1/2/3 \u6b21\u6709\u6548\u91c7\u96c6\u3002",
        "3. \u6709\u6548\u91c7\u96c6\u95f4\u9694\u9700\u5927\u4e8e\u7b49\u4e8e 30 \u5206\u949f\uff1b30 \u5206\u949f\u5185\u8bb0\u5f55\u89c6\u4e3a\u540c\u4e00\u5019\u9009\u65f6\u6bb5\u3002",
        "4. \u540c\u65e5\u540c\u69fd\u4f4d\u7684\u4e2d\u79d1/\u7389\u751f\u5802\u8bb0\u5f55\u5408\u5e76\u4e3a\u4e00\u4e2a visit \u7684\u4e0d\u540c\u6a21\u6001\u3002",
        "5. \u65f6\u95f4\u91cd\u53e0\u65f6\uff0c\u9009\u62e9\u6a21\u6001\u6700\u5b8c\u6574\u7684\u4e00\u6b21\u4f5c\u4e3a\u6709\u6548\u4ee3\u8868\u3002",
        "6. \u4fdd\u7559\u59d3\u540d\u4e0d\u4e00\u81f4\u3001\u7591\u4f3c\u91cd\u590d/\u9ad8\u5ea6\u76f8\u4f3c\u6570\u636e\u300110 \u5206\u949f\u5185\u5f02\u5e38\u6253\u5361\u7b49\u6807\u8bb0\u3002",
    ])
    slide4 = slide_xml([
        rect(2, 0, 0, SLIDE_W, SLIDE_H, "F8FAFC", "F8FAFC"),
        text_box(3, emu(0.65), emu(0.45), emu(12.0), emu(0.55), "\u7edf\u8ba1\u89c4\u5219", 29, True, "0F172A"),
        text_box(4, emu(0.85), emu(1.25), emu(11.4), emu(4.65), rules, 19, False, "334155"),
        text_box(5, emu(0.85), emu(6.15), emu(11.2), emu(0.45), f"\u51bb\u7ed3\u89c4\u5219\u7248\u672c\uff1a{RULE_VERSION}", 17, True, "0F766E"),
    ])

    slide5 = slide_xml([
        rect(2, 0, 0, SLIDE_W, SLIDE_H, "FFFFFF", "FFFFFF"),
        text_box(3, emu(0.65), emu(0.45), emu(12.0), emu(0.55), "\u521d\u6b65\u89e3\u8bfb", 29, True, "0F172A"),
        text_box(4, emu(0.85), emu(1.25), emu(11.3), emu(4.9), "\n".join([
            f"\u603b\u8bb0\u5f55 {total} \u6761\uff0c\u5b8c\u6574\u8bb0\u5f55 {status.get('complete', 0)} \u6761\uff0c\u5b8c\u6574\u7387 {complete_rate:.1%}\u3002",
            f"\u4e0d\u5b8c\u6574\u8bb0\u5f55 {status.get('incomplete', 0)} \u6761\uff0c\u5e94\u5728\u5165\u5e93\u65f6\u4fdd\u7559\u7f3a\u5931\u6a21\u6001\u660e\u7ec6\u3002",
            f"\u7591\u4f3c\u5f02\u5e38 {status.get('suspicious', 0)} \u6761\uff0c\u65e0\u6548 {status.get('invalid', 0)} \u6761\uff0c\u540e\u7eed\u5206\u6790\u5efa\u8bae\u4f5c\u4e3a\u8d28\u63a7\u6807\u8bb0\u800c\u975e\u76f4\u63a5\u5220\u9664\u3002",
            f"\u59d3\u540d\u4e0d\u4e00\u81f4\u6807\u8bb0 {stats['mismatch']} \u6761\uff0c\u9700\u7ee7\u7eed\u4f9d\u8d56\u59d3\u540d\u6620\u5c04\u914d\u7f6e\u548c\u624b\u5de5\u590d\u6838\u72b6\u6001\u3002",
        ]), 21, False, "334155"),
        text_box(5, emu(0.85), emu(6.25), emu(11.5), emu(0.45), "\u672c PPT \u57fa\u4e8e\u5df2\u751f\u6210\u7684 Excel \u548c PNG \u77e9\u9635\u56fe\u81ea\u52a8\u751f\u6210\u3002", 15, False, "64748B"),
    ])
    return [(slide1, None), (slide2, None), (slide3, "image"), (slide4, None), (slide5, None)]


def write_pptx(slides: list[tuple[str, str | None]]) -> None:
    slide_overrides = "\n".join(
        f'<Override PartName="/ppt/slides/slide{i}.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.slide+xml"/>'
        for i in range(1, len(slides) + 1)
    )
    content_types = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Default Extension="png" ContentType="image/png"/>
  <Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>
  <Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>
  <Override PartName="/ppt/presentation.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.presentation.main+xml"/>
  <Override PartName="/ppt/theme/theme1.xml" ContentType="application/vnd.openxmlformats-officedocument.theme+xml"/>
  <Override PartName="/ppt/slideMasters/slideMaster1.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.slideMaster+xml"/>
  <Override PartName="/ppt/slideLayouts/slideLayout1.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.slideLayout+xml"/>
  {slide_overrides}
</Types>"""

    sld_ids = "\n".join(f'<p:sldId id="{255 + i}" r:id="rId{i}"/>' for i in range(1, len(slides) + 1))
    presentation = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<p:presentation xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
 xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main">
 <p:sldMasterIdLst><p:sldMasterId id="2147483648" r:id="rId{len(slides)+1}"/></p:sldMasterIdLst>
 <p:sldIdLst>{sld_ids}</p:sldIdLst>
 <p:sldSz cx="{SLIDE_W}" cy="{SLIDE_H}" type="wide"/>
 <p:notesSz cx="6858000" cy="9144000"/>
</p:presentation>"""

    pres_rels = [(f"rId{i}", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide", f"slides/slide{i}.xml") for i in range(1, len(slides) + 1)]
    pres_rels.append((f"rId{len(slides)+1}", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/slideMaster", "slideMasters/slideMaster1.xml"))
    pres_rels.append((f"rId{len(slides)+2}", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/theme", "theme/theme1.xml"))

    master = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<p:sldMaster xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main">
<p:cSld><p:spTree><p:nvGrpSpPr><p:cNvPr id="1" name=""/><p:cNvGrpSpPr/><p:nvPr/></p:nvGrpSpPr><p:grpSpPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/><a:chOff x="0" y="0"/><a:chExt cx="0" cy="0"/></a:xfrm></p:grpSpPr></p:spTree></p:cSld>
<p:sldLayoutIdLst><p:sldLayoutId id="2147483649" r:id="rId1"/></p:sldLayoutIdLst><p:txStyles><p:titleStyle/><p:bodyStyle/><p:otherStyle/></p:txStyles></p:sldMaster>"""
    layout = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<p:sldLayout xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main" type="blank" preserve="1"><p:cSld name="Blank"><p:spTree><p:nvGrpSpPr><p:cNvPr id="1" name=""/><p:cNvGrpSpPr/><p:nvPr/></p:nvGrpSpPr><p:grpSpPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/><a:chOff x="0" y="0"/><a:chExt cx="0" cy="0"/></a:xfrm></p:grpSpPr></p:spTree></p:cSld><p:clrMapOvr><a:masterClrMapping/></p:clrMapOvr></p:sldLayout>"""
    theme = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?><a:theme xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" name="TCM"><a:themeElements><a:clrScheme name="TCM"><a:dk1><a:srgbClr val="0F172A"/></a:dk1><a:lt1><a:srgbClr val="FFFFFF"/></a:lt1><a:dk2><a:srgbClr val="334155"/></a:dk2><a:lt2><a:srgbClr val="F8FAFC"/></a:lt2><a:accent1><a:srgbClr val="0F766E"/></a:accent1><a:accent2><a:srgbClr val="2563EB"/></a:accent2><a:accent3><a:srgbClr val="D97706"/></a:accent3><a:accent4><a:srgbClr val="BE123C"/></a:accent4><a:accent5><a:srgbClr val="7C3AED"/></a:accent5><a:accent6><a:srgbClr val="0891B2"/></a:accent6><a:hlink><a:srgbClr val="2563EB"/></a:hlink><a:folHlink><a:srgbClr val="7C3AED"/></a:folHlink></a:clrScheme><a:fontScheme name="TCM"><a:majorFont><a:latin typeface="Microsoft YaHei"/><a:ea typeface="Microsoft YaHei"/></a:majorFont><a:minorFont><a:latin typeface="Microsoft YaHei"/><a:ea typeface="Microsoft YaHei"/></a:minorFont></a:fontScheme><a:fmtScheme name="TCM"><a:fillStyleLst/><a:lnStyleLst/><a:effectStyleLst/><a:bgFillStyleLst/></a:fmtScheme></a:themeElements></a:theme>"""
    core = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?><cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"><dc:title>TCM Check-in Summary</dc:title><dc:creator>Codex</dc:creator><dcterms:created xsi:type="dcterms:W3CDTF">{datetime.now().isoformat()}</dcterms:created></cp:coreProperties>"""
    app = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties"><Application>Codex</Application></Properties>"""

    with zipfile.ZipFile(OUTPUT, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", content_types)
        z.writestr("_rels/.rels", rels_xml([("rId1", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument", "ppt/presentation.xml"), ("rId2", "http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties", "docProps/core.xml"), ("rId3", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties", "docProps/app.xml")]))
        z.writestr("docProps/core.xml", core)
        z.writestr("docProps/app.xml", app)
        z.writestr("ppt/presentation.xml", presentation)
        z.writestr("ppt/_rels/presentation.xml.rels", rels_xml(pres_rels))
        z.writestr("ppt/slideMasters/slideMaster1.xml", master)
        z.writestr("ppt/slideMasters/_rels/slideMaster1.xml.rels", rels_xml([("rId1", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/slideLayout", "../slideLayouts/slideLayout1.xml")]))
        z.writestr("ppt/slideLayouts/slideLayout1.xml", layout)
        z.writestr("ppt/slideLayouts/_rels/slideLayout1.xml.rels", rels_xml([("rId1", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/slideMaster", "../slideMasters/slideMaster1.xml")]))
        z.writestr("ppt/theme/theme1.xml", theme)
        z.write(HEATMAP, "ppt/media/matrix.png")
        for i, (xml, image_kind) in enumerate(slides, start=1):
            z.writestr(f"ppt/slides/slide{i}.xml", xml)
            rels = [("rId1", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/slideLayout", "../slideLayouts/slideLayout1.xml")]
            if image_kind == "image":
                rels.append(("rId2", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image", "../media/matrix.png"))
            z.writestr(f"ppt/slides/_rels/slide{i}.xml.rels", rels_xml(rels))


def main() -> None:
    stats = load_stats()
    slides = build_slides(stats)
    write_pptx(slides)
    print(OUTPUT)


if __name__ == "__main__":
    main()
